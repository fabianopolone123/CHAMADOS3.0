import json
import os
import tempfile
from decimal import Decimal

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from .models import (
    AtendimentoHistorico,
    Chamado,
    ChamadoEvento,
    ChamadoMensagem,
    ChamadoMensagemAnexo,
    CofreAuditoria,
    CofreConfig,
    CofreCredencial,
    ContaEmail,
    Contrato,
    ContratoAnexo,
    Dica,
    EmprestimoTI,
    EquipamentoEmprestimoTI,
    EnderecoIP,
    FuturaDigital,
    ItemMenuConfig,
    Licenca,
    LicencaSoftware,
    OrcamentoContrato,
    OrcamentoDocumento,
    PainelAuditoria,
    PausaAutomatica,
    PendenciaTI,
    Ramal,
    RequisicaoContrato,
    RequisicaoContratoEvento,
    ServicoFeito,
    ServicoFeitoAnexo,
    Starlink,
    SuborcamentoContrato,
    SuborcamentoDocumento,
)
from . import painel_dados
from .menu import CHAVES_PADRAO, itens_menu_para_painel
from .permissions import ADMIN_GROUP_NAME, ATTENDANT_GROUP_NAME


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class ChamadoMensagemTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.owner = User.objects.create_user(username="joao", password="x")
        self.other = User.objects.create_user(username="maria", password="x")
        self.attendant = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.attendant.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

        self.chamado = Chamado.objects.create(
            numero="CH-000001",
            titulo="Impressora sem tinta",
            descricao="Nao imprime nada",
            solicitante=self.owner,
            solicitante_nome="Joao",
            status=Chamado.STATUS_ABERTO,
        )

    def _post_message(self, data):
        return self.client.post(
            reverse("ticket_message_create", args=[self.chamado.numero]), data
        )

    def test_owner_sends_message_without_attachment(self):
        self.client.force_login(self.owner)
        resp = self._post_message({"texto": "Bom dia, alguma novidade?"})
        self.assertEqual(resp.status_code, 302)

        mensagem = ChamadoMensagem.objects.get(chamado=self.chamado)
        self.assertEqual(mensagem.texto, "Bom dia, alguma novidade?")
        self.assertEqual(mensagem.autor, self.owner)

        # historico tecnico guarda apenas o resumo, nao o texto da mensagem
        evento = ChamadoEvento.objects.get(chamado=self.chamado, tipo=ChamadoEvento.TIPO_COMENTARIO)
        self.assertIn("Mensagem enviada pelo solicitante", evento.descricao)
        self.assertNotIn("Bom dia", evento.descricao)

    def test_owner_sends_message_with_attachment(self):
        self.client.force_login(self.owner)
        arquivo = SimpleUploadedFile("erro.txt", b"conteudo", content_type="text/plain")
        resp = self._post_message({"texto": "Segue print", "anexos": [arquivo]})
        self.assertEqual(resp.status_code, 302)

        mensagem = ChamadoMensagem.objects.get(chamado=self.chamado)
        self.assertEqual(mensagem.anexos.count(), 1)

        evento = ChamadoEvento.objects.get(chamado=self.chamado, tipo=ChamadoEvento.TIPO_COMENTARIO)
        self.assertIn("1 anexo(s)", evento.descricao)

    def test_empty_message_is_rejected(self):
        self.client.force_login(self.owner)
        resp = self._post_message({"texto": "   "})
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(ChamadoMensagem.objects.filter(chamado=self.chamado).exists())

    def test_other_common_user_cannot_send_message(self):
        self.client.force_login(self.other)
        resp = self._post_message({"texto": "invadindo"})
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(ChamadoMensagem.objects.filter(chamado=self.chamado).exists())

    def test_attendant_can_reply_on_any_ticket(self):
        self.client.force_login(self.attendant)
        resp = self._post_message({"texto": "Estamos verificando."})
        self.assertEqual(resp.status_code, 302)

        mensagem = ChamadoMensagem.objects.get(chamado=self.chamado)
        self.assertEqual(mensagem.autor, self.attendant)
        evento = ChamadoEvento.objects.get(chamado=self.chamado, tipo=ChamadoEvento.TIPO_COMENTARIO)
        self.assertIn("Mensagem enviada por", evento.descricao)
        self.assertNotIn("solicitante", evento.descricao)

    def test_other_common_user_cannot_access_detail(self):
        self.client.force_login(self.other)
        resp = self.client.get(reverse("ticket_detail", args=[self.chamado.numero]))
        self.assertEqual(resp.status_code, 302)  # redirecionado, sem acesso

    def test_message_attachment_download_permission(self):
        self.client.force_login(self.owner)
        arquivo = SimpleUploadedFile("nota.txt", b"abc", content_type="text/plain")
        self._post_message({"texto": "arquivo", "anexos": [arquivo]})
        anexo = ChamadoMensagemAnexo.objects.get()

        url = reverse("download_message_anexo", args=[self.chamado.numero, anexo.id])
        self.assertEqual(self.client.get(url).status_code, 200)

        self.client.force_login(self.other)
        self.assertEqual(self.client.get(url).status_code, 404)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class EncerramentoChamadoTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.owner = User.objects.create_user(username="joao", password="x")
        self.attendant = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.attendant.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

        self.chamado = Chamado.objects.create(
            numero="CH-000010",
            titulo="Rede caiu",
            descricao="Sem internet no setor",
            solicitante=self.owner,
            status=Chamado.STATUS_EM_ATENDIMENTO,
            atendente_atual=self.attendant,
        )

    def _start_attendance(self, user):
        return AtendimentoHistorico.objects.create(
            chamado=self.chamado, atendente=user, iniciado_em=timezone.now()
        )

    def _finish(self, action, description="Feito"):
        return self.client.post(
            reverse("finish_attendance"),
            data=json.dumps(
                {"ticket_number": self.chamado.numero, "action": action, "description": description}
            ),
            content_type="application/json",
        )

    def test_stop_closes_and_moves_ticket(self):
        self.client.force_login(self.attendant)
        self._start_attendance(self.attendant)
        resp = self._finish("stop", "Resolvido e testado com o usuario")
        self.assertEqual(resp.status_code, 200)

        data = resp.json()
        self.assertTrue(data["ticket_closed"])
        self.assertEqual(data["status"], Chamado.STATUS_FECHADO)
        self.assertEqual(data["status_class"], "status-neutral")

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_FECHADO)
        self.assertIsNotNone(self.chamado.fechado_em)
        self.assertEqual(self.chamado.atendente_atual, self.attendant)

        self.assertTrue(
            ChamadoEvento.objects.filter(chamado=self.chamado, tipo=ChamadoEvento.TIPO_STATUS).exists()
        )
        # O historico tecnico registra a finalizacao com quem finalizou e o texto
        # de "O que foi feito" (registro tecnico, separado da conversa do usuario).
        self.assertTrue(
            ChamadoEvento.objects.filter(
                chamado=self.chamado, descricao__icontains="finalizado"
            ).exists()
        )
        self.assertTrue(
            ChamadoEvento.objects.filter(
                chamado=self.chamado,
                descricao__icontains="O que foi feito: Resolvido e testado com o usuario",
            ).exists()
        )

    def test_pause_does_not_close_ticket(self):
        self.client.force_login(self.attendant)
        self._start_attendance(self.attendant)
        resp = self._finish("pause", "Aguardando peca de reposicao")
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(resp.json().get("ticket_closed"))

        # Pause sem motivo encerra o periodo de atendimento e devolve o chamado
        # para "Atribuido" (nao fica "Em atendimento" sem Play ativo).
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_ATRIBUIDO)

    def _por_em_aguardando(self, status=Chamado.STATUS_AGUARDANDO_PECA):
        self.chamado.status = status
        self.chamado.save(update_fields=["status"])

    def test_stop_sem_play_fecha_chamado_em_aguardando(self):
        # Chamado parado em "aguardando peca" pode ser encerrado direto pelo Stop,
        # sem precisar iniciar um atendimento so para fechar.
        self._por_em_aguardando()
        self.client.force_login(self.attendant)
        resp = self._finish("stop", "Peca nao veio mais, usuario trocou de maquina")
        self.assertEqual(resp.status_code, 200)

        data = resp.json()
        self.assertTrue(data["ticket_closed"])
        self.assertEqual(data["status"], Chamado.STATUS_FECHADO)

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_FECHADO)
        self.assertIsNotNone(self.chamado.fechado_em)
        # Nenhum periodo de atendimento e inventado para o fechamento.
        self.assertEqual(AtendimentoHistorico.objects.filter(chamado=self.chamado).count(), 0)

    def test_stop_sem_play_registra_tudo_na_linha_do_tempo(self):
        self._por_em_aguardando(Chamado.STATUS_AGUARDANDO_USUARIO)
        self.client.force_login(self.attendant)
        self._finish("stop", "Usuario nao respondeu, chamado encerrado")

        # Mudanca de status (de onde saiu) + encerramento com autor e o que foi feito.
        self.assertTrue(
            ChamadoEvento.objects.filter(
                chamado=self.chamado,
                tipo=ChamadoEvento.TIPO_STATUS,
                descricao__icontains="Aguardando usuario",
            ).exists()
        )
        encerramento = ChamadoEvento.objects.filter(
            chamado=self.chamado, tipo=ChamadoEvento.TIPO_ENCERRAMENTO_DIRETO
        ).first()
        self.assertIsNotNone(encerramento)
        self.assertIn("sem atendimento ativo", encerramento.descricao)
        self.assertIn("Aguardando usuario", encerramento.descricao)
        self.assertIn("O que foi feito: Usuario nao respondeu", encerramento.descricao)

        # O encerramento tambem aparece no andamento do chamado (detalhe).
        resp = self.client.get(reverse("ticket_detail", args=[self.chamado.numero]))
        acoes = [item["action"] for item in resp.context["timeline"]]
        self.assertIn("encerrou o chamado sem atendimento ativo", acoes)

    def test_stop_sem_play_fora_de_aguardando_e_bloqueado(self):
        # Sem Play e sem estar em "aguardando", continua valendo a regra antiga.
        self.chamado.status = Chamado.STATUS_ATRIBUIDO
        self.chamado.save(update_fields=["status"])
        self.client.force_login(self.attendant)
        resp = self._finish("stop", "tentando fechar sem play")
        self.assertEqual(resp.status_code, 409)

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_ATRIBUIDO)

    def test_pause_sem_play_continua_bloqueado(self):
        self._por_em_aguardando()
        self.client.force_login(self.attendant)
        resp = self._finish("pause", "tentando pausar sem play")
        self.assertEqual(resp.status_code, 409)

    def test_stop_sem_play_exige_descricao(self):
        self._por_em_aguardando()
        self.client.force_login(self.attendant)
        resp = self._finish("stop", "")
        self.assertEqual(resp.status_code, 400)

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_AGUARDANDO_PECA)

    def test_stop_sem_play_em_chamado_ja_fechado_e_bloqueado(self):
        self.chamado.status = Chamado.STATUS_FECHADO
        self.chamado.save(update_fields=["status"])
        self.client.force_login(self.attendant)
        resp = self._finish("stop", "fechando de novo")
        self.assertEqual(resp.status_code, 409)

    def test_usuario_comum_nao_fecha_chamado_em_aguardando(self):
        self._por_em_aguardando()
        self.client.force_login(self.owner)
        resp = self._finish("stop", "tentando encerrar")
        self.assertEqual(resp.status_code, 403)

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_AGUARDANDO_PECA)

    def test_card_em_aguardando_libera_stop_sem_play(self):
        # O card so mostra o Stop sem Play quando esta em "aguardando".
        self._por_em_aguardando()
        self.client.force_login(self.attendant)
        coluna = next(
            c
            for c in self.client.get(reverse("tickets_dashboard")).context["attendant_columns"]
            if c["attendant_id"] == self.attendant.id
        )
        card = next(c for c in coluna["tickets"] if c["number"] == self.chamado.numero)
        self.assertTrue(card["can_close_direct"])

        # Com Play ativo o fluxo normal (Pause/Stop do atendimento) e que vale.
        self._start_attendance(self.attendant)
        coluna = next(
            c
            for c in self.client.get(reverse("tickets_dashboard")).context["attendant_columns"]
            if c["attendant_id"] == self.attendant.id
        )
        card = next(c for c in coluna["tickets"] if c["number"] == self.chamado.numero)
        self.assertFalse(card["can_close_direct"])

    def test_common_user_cannot_finish_attendance(self):
        self.client.force_login(self.owner)
        resp = self._finish("stop", "tentando encerrar")
        self.assertEqual(resp.status_code, 403)

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_EM_ATENDIMENTO)

    def test_drag_to_fechado_is_blocked(self):
        # O fechamento so acontece via Stop: o endpoint de movimentacao recusa o
        # destino "fechado" e nao altera o chamado.
        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("move_ticket"),
            data=json.dumps({"ticket_number": self.chamado.numero, "target": "fechado"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 409)

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_EM_ATENDIMENTO)
        self.assertIsNone(self.chamado.fechado_em)

    def test_dashboard_stats_por_atendente(self):
        # A coluna do atendente traz a quebra por status: em atendimento,
        # aguardando (soma dos 3) e atribuido. self.chamado esta "em_atendimento"
        # com Play ativo (o status so conta como em atendimento com Play).
        self._start_attendance(self.attendant)
        Chamado.objects.create(
            numero="CH-000020", titulo="Peca", solicitante=self.owner,
            status=Chamado.STATUS_AGUARDANDO_PECA, atendente_atual=self.attendant,
        )
        Chamado.objects.create(
            numero="CH-000021", titulo="Usuario", solicitante=self.owner,
            status=Chamado.STATUS_AGUARDANDO_USUARIO, atendente_atual=self.attendant,
        )
        Chamado.objects.create(
            numero="CH-000022", titulo="Atribuido", solicitante=self.owner,
            status=Chamado.STATUS_ATRIBUIDO, atendente_atual=self.attendant,
        )
        self.client.force_login(self.attendant)
        resp = self.client.get(reverse("tickets_dashboard"))
        self.assertEqual(resp.status_code, 200)
        coluna = next(
            c for c in resp.context["attendant_columns"] if c["attendant_id"] == self.attendant.id
        )
        self.assertEqual(coluna["count"], 4)
        self.assertEqual(coluna["stats"]["em_atendimento"], 1)
        self.assertEqual(coluna["stats"]["aguardando"], 2)
        self.assertEqual(coluna["stats"]["atribuido"], 1)

    def test_status_em_atendimento_sem_play_conta_como_atribuido(self):
        # Dado antigo: chamado gravado como "em_atendimento" mas sem nenhum
        # atendimento ativo (antes, arrastar para o atendente ja marcava assim).
        # Ele nao pode inflar o contador "em atend." da coluna nem o badge.
        Chamado.objects.create(
            numero="CH-000023", titulo="Legado sem play", solicitante=self.owner,
            status=Chamado.STATUS_EM_ATENDIMENTO, atendente_atual=self.attendant,
        )
        self._start_attendance(self.attendant)  # so self.chamado esta em play

        self.client.force_login(self.attendant)
        resp = self.client.get(reverse("tickets_dashboard"))
        coluna = next(
            c for c in resp.context["attendant_columns"] if c["attendant_id"] == self.attendant.id
        )
        self.assertEqual(coluna["count"], 2)
        self.assertEqual(coluna["stats"]["em_atendimento"], 1)  # so o que tem Play
        self.assertEqual(coluna["stats"]["atribuido"], 1)

        card = next(c for c in coluna["tickets"] if c["number"] == "CH-000023")
        self.assertEqual(card["status"], Chamado.STATUS_ATRIBUIDO)
        self.assertEqual(card["status_label"], "Atribuido")

    def test_status_em_atendimento_sem_play_e_sem_atendente_conta_como_aberto(self):
        # Mesmo caso, mas sem atendente atual: o chamado vive na coluna
        # "Chamados abertos" e deve aparecer como Aberto.
        self.chamado.atendente_atual = None
        self.chamado.save(update_fields=["atendente_atual"])

        self.client.force_login(self.attendant)
        resp = self.client.get(reverse("tickets_dashboard"))
        card = next(
            c for c in resp.context["open_column"]["tickets"] if c["number"] == self.chamado.numero
        )
        self.assertEqual(card["status"], Chamado.STATUS_ABERTO)
        self.assertEqual(card["status_label"], "Aberto")

    def test_mover_chamado_com_play_ativo_e_bloqueado(self):
        # Com um atendimento ativo (Play), o chamado nao pode ser movido: o
        # endpoint recusa com 409 e nada muda (atendente/status preservados).
        User = get_user_model()
        outro = User.objects.create_user(username="ti2", password="x")
        outro.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))
        self._start_attendance(self.attendant)  # Play ativo do self.attendant

        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("move_ticket"),
            data=json.dumps(
                {"ticket_number": self.chamado.numero, "target": "atendente", "attendant_id": outro.id}
            ),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 409)

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.atendente_atual, self.attendant)  # nao mudou
        self.assertEqual(self.chamado.status, Chamado.STATUS_EM_ATENDIMENTO)
        # O Play continua sendo do atendente original.
        self.assertTrue(
            AtendimentoHistorico.objects.filter(
                chamado=self.chamado, atendente=self.attendant, finalizado_em__isnull=True
            ).exists()
        )

    def test_mover_para_abertos_com_play_ativo_e_bloqueado(self):
        # A regra vale para qualquer destino: nem devolver para "Chamados abertos"
        # e permitido enquanto ha Play ativo.
        self._start_attendance(self.attendant)
        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("move_ticket"),
            data=json.dumps({"ticket_number": self.chamado.numero, "target": "aberto"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 409)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.atendente_atual, self.attendant)

    def test_move_liberado_apos_pausar_o_atendimento(self):
        # Depois de pausar o Play, o chamado volta a poder ser movido.
        User = get_user_model()
        outro = User.objects.create_user(username="ti3", password="x")
        outro.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))
        self._start_attendance(self.attendant)
        self.client.force_login(self.attendant)
        self.assertEqual(self._finish("pause", "pausando para repassar").status_code, 200)

        resp = self.client.post(
            reverse("move_ticket"),
            data=json.dumps(
                {"ticket_number": self.chamado.numero, "target": "atendente", "attendant_id": outro.id}
            ),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.atendente_atual, outro)
        self.assertEqual(self.chamado.status, Chamado.STATUS_ATRIBUIDO)

    def test_stop_requires_active_attendance(self):
        # Sem Play ativo, o Stop e recusado e o chamado nao e fechado.
        self.client.force_login(self.attendant)
        resp = self._finish("stop", "sem play ativo")
        self.assertEqual(resp.status_code, 409)

        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_EM_ATENDIMENTO)

    def test_play_em_varios_chamados_ao_mesmo_tempo(self):
        # Multiplos atendimentos ativos sao permitidos: dar Play em dois chamados
        # mantem os dois ativos (nao pausa nem bloqueia).
        self.client.force_login(self.attendant)
        ch2 = Chamado.objects.create(
            numero="CH-000011", titulo="Outro chamado", solicitante=self.owner,
            status=Chamado.STATUS_EM_ATENDIMENTO, atendente_atual=self.attendant,
        )

        def _play(numero):
            return self.client.post(
                reverse("start_attendance"),
                data=json.dumps({"ticket_number": numero}),
                content_type="application/json",
            )

        self.assertEqual(_play(self.chamado.numero).status_code, 200)
        self.assertEqual(_play(ch2.numero).status_code, 200)

        ativos = AtendimentoHistorico.objects.filter(
            atendente=self.attendant, finalizado_em__isnull=True
        )
        self.assertEqual(ativos.count(), 2)  # os dois ativos ao mesmo tempo

        # Play repetido no MESMO chamado ainda e bloqueado (nao duplica).
        self.assertEqual(_play(ch2.numero).status_code, 409)

    def test_stop_age_no_chamado_especifico(self):
        # Com varios ativos, o Stop/Pause encerra o atendimento do chamado informado.
        self.client.force_login(self.attendant)
        ch2 = Chamado.objects.create(
            numero="CH-000012", titulo="Segundo", solicitante=self.owner,
            status=Chamado.STATUS_EM_ATENDIMENTO, atendente_atual=self.attendant,
        )
        self._start_attendance(self.attendant)  # ativo no self.chamado
        AtendimentoHistorico.objects.create(chamado=ch2, atendente=self.attendant, iniciado_em=timezone.now())

        resp = self.client.post(
            reverse("finish_attendance"),
            data=json.dumps({"ticket_number": ch2.numero, "action": "pause", "description": "pausando o 2"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        # so o ch2 foi finalizado; o self.chamado continua ativo
        self.assertFalse(
            AtendimentoHistorico.objects.filter(chamado=ch2, finalizado_em__isnull=True).exists()
        )
        self.assertTrue(
            AtendimentoHistorico.objects.filter(chamado=self.chamado, finalizado_em__isnull=True).exists()
        )

    def test_pause_com_motivo_marca_aguardando_e_registra_historico(self):
        self.client.force_login(self.attendant)
        self._start_attendance(self.attendant)
        resp = self.client.post(
            reverse("finish_attendance"),
            data=json.dumps({
                "ticket_number": self.chamado.numero, "action": "pause",
                "description": "peca pedida ao fornecedor", "pause_reason": "aguardando_peca",
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_AGUARDANDO_PECA)
        # pausa registrada no historico do chamado
        self.assertTrue(
            ChamadoEvento.objects.filter(chamado=self.chamado, descricao__icontains="pausado").exists()
        )

    def test_move_para_atendente_marca_atribuido_nao_em_atendimento(self):
        # Arrastar para um atendente apenas atribui o chamado; "Em atendimento"
        # so vale com Play ativo. Depois o Play muda o status para em_atendimento.
        self.chamado.status = Chamado.STATUS_ABERTO
        self.chamado.atendente_atual = None
        self.chamado.save()
        self.client.force_login(self.attendant)

        resp = self.client.post(
            reverse("move_ticket"),
            data=json.dumps(
                {"ticket_number": self.chamado.numero, "target": "atendente", "attendant_id": self.attendant.id}
            ),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["status"], Chamado.STATUS_ATRIBUIDO)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_ATRIBUIDO)
        self.assertFalse(
            AtendimentoHistorico.objects.filter(chamado=self.chamado, finalizado_em__isnull=True).exists()
        )

        # Play inicia o periodo de atendimento e move para "Em atendimento".
        resp = self.client.post(
            reverse("start_attendance"),
            data=json.dumps({"ticket_number": self.chamado.numero}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_EM_ATENDIMENTO)

    def test_play_registra_historico_e_retoma_em_atendimento(self):
        self.chamado.status = Chamado.STATUS_AGUARDANDO_PECA
        self.chamado.save()
        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("start_attendance"),
            data=json.dumps({"ticket_number": self.chamado.numero}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, Chamado.STATUS_EM_ATENDIMENTO)
        self.assertTrue(
            ChamadoEvento.objects.filter(
                chamado=self.chamado, descricao__icontains="Atendimento iniciado"
            ).exists()
        )


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class PendenciaTITests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.creator = User.objects.create_user(username="ti1", password="x")
        self.attendant = User.objects.create_user(username="ti2", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        grupo = Group.objects.get(name=ATTENDANT_GROUP_NAME)
        self.creator.groups.add(grupo)
        self.attendant.groups.add(grupo)

    def _create_pendencia(self, autor):
        return PendenciaTI.objects.create(
            titulo="Trocar switch", descricao="Switch do 2o andar", criado_por=autor
        )

    def test_attendant_creates_pendencia(self):
        self.client.force_login(self.creator)
        resp = self.client.post(
            reverse("pendencia_create"),
            data=json.dumps({"titulo": "Comprar toner", "descricao": "Impressora sala 3"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(PendenciaTI.objects.filter(titulo="Comprar toner").exists())

    def test_common_user_cannot_create_or_view_pendencia(self):
        self.client.force_login(self.common)
        resp = self.client.post(
            reverse("pendencia_create"),
            data=json.dumps({"titulo": "Invadindo", "descricao": "nao pode"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

        pend = self._create_pendencia(self.creator)
        resp = self.client.get(reverse("pendencia_detail", args=[pend.id]))
        self.assertEqual(resp.status_code, 403)

    def test_convert_pendencia_creates_chamado(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("pendencia_convert", args=[pend.id]),
            data=json.dumps({"attendant_id": self.attendant.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()

        chamado = Chamado.objects.get(numero=data["ticket_number"])
        self.assertEqual(chamado.titulo, "Trocar switch")
        self.assertEqual(chamado.solicitante, self.creator)  # solicitante = quem criou a pendencia
        self.assertEqual(chamado.atendente_atual, self.attendant)  # atendente da coluna destino
        # Convertida para a coluna do atendente = atribuida (sem Play ativo ainda).
        self.assertEqual(chamado.status, Chamado.STATUS_ATRIBUIDO)

        pend.refresh_from_db()
        self.assertTrue(pend.convertido_em_chamado)
        self.assertEqual(pend.chamado_gerado, chamado)
        self.assertEqual(pend.convertido_por, self.attendant)

        self.assertTrue(
            ChamadoEvento.objects.filter(chamado=chamado, descricao__icontains="pendencia").exists()
        )

    def test_convert_twice_does_not_duplicate(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.attendant)
        url = reverse("pendencia_convert", args=[pend.id])
        body = json.dumps({"attendant_id": self.attendant.id})

        first = self.client.post(url, data=body, content_type="application/json")
        self.assertEqual(first.status_code, 200)
        second = self.client.post(url, data=body, content_type="application/json")
        self.assertEqual(second.status_code, 409)

        self.assertEqual(Chamado.objects.filter(pendencias_origem=pend).count(), 1)

    def test_convert_rejects_non_attendant_target(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("pendencia_convert", args=[pend.id]),
            data=json.dumps({"attendant_id": self.common.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)
        pend.refresh_from_db()
        self.assertFalse(pend.convertido_em_chamado)

    def test_common_user_cannot_convert(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.common)
        resp = self.client.post(
            reverse("pendencia_convert", args=[pend.id]),
            data=json.dumps({"attendant_id": self.attendant.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_attendant_deletes_pendencia(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.attendant)
        resp = self.client.post(reverse("pendencia_delete", args=[pend.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(PendenciaTI.objects.filter(id=pend.id).exists())

    def test_common_user_cannot_delete_pendencia(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.common)
        resp = self.client.post(reverse("pendencia_delete", args=[pend.id]))
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(PendenciaTI.objects.filter(id=pend.id).exists())

    def test_delete_pendencia_requires_post(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.attendant)
        resp = self.client.get(reverse("pendencia_delete", args=[pend.id]))
        self.assertEqual(resp.status_code, 405)

    def test_create_pendencia_without_priority_is_colorless(self):
        self.client.force_login(self.creator)
        resp = self.client.post(
            reverse("pendencia_create"),
            data=json.dumps({"titulo": "Sem prioridade", "descricao": "sem cor"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        pend = PendenciaTI.objects.get(titulo="Sem prioridade")
        self.assertIsNone(pend.prioridade)
        self.assertEqual(pend.cor, "")

    def test_create_pendencia_with_priority(self):
        self.client.force_login(self.creator)
        resp = self.client.post(
            reverse("pendencia_create"),
            data=json.dumps(
                {"titulo": "Urgente", "descricao": "vermelho", "prioridade": 1}
            ),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        pend = PendenciaTI.objects.get(titulo="Urgente")
        self.assertEqual(pend.prioridade, 1)

    def test_create_pendencia_invalid_priority_is_colorless(self):
        self.client.force_login(self.creator)
        resp = self.client.post(
            reverse("pendencia_create"),
            data=json.dumps(
                {"titulo": "Invalida", "descricao": "fora da escala", "prioridade": 99}
            ),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        pend = PendenciaTI.objects.get(titulo="Invalida")
        self.assertIsNone(pend.prioridade)

    def test_detail_returns_priority_and_color(self):
        pend = self._create_pendencia(self.creator)
        pend.prioridade = 2
        pend.save(update_fields=["prioridade"])
        self.client.force_login(self.attendant)
        resp = self.client.get(reverse("pendencia_detail", args=[pend.id]))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["prioridade"], 2)
        self.assertEqual(data["cor"], PendenciaTI.PRIORIDADE_CORES[2])

    def test_update_priority(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("pendencia_priority", args=[pend.id]),
            data=json.dumps({"prioridade": 5}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        pend.refresh_from_db()
        self.assertEqual(pend.prioridade, 5)
        self.assertEqual(resp.json()["cor"], PendenciaTI.PRIORIDADE_CORES[5])

    def test_update_priority_requires_post(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.attendant)
        resp = self.client.get(reverse("pendencia_priority", args=[pend.id]))
        self.assertEqual(resp.status_code, 405)

    def test_common_user_cannot_update_priority(self):
        pend = self._create_pendencia(self.creator)
        self.client.force_login(self.common)
        resp = self.client.post(
            reverse("pendencia_priority", args=[pend.id]),
            data=json.dumps({"prioridade": 1}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_priority_ordering_red_first_colorless_last(self):
        sem_cor = PendenciaTI.objects.create(
            titulo="Sem cor", descricao="branco", criado_por=self.creator
        )
        baixa = PendenciaTI.objects.create(
            titulo="Baixa", descricao="verde", criado_por=self.creator, prioridade=5
        )
        urgente = PendenciaTI.objects.create(
            titulo="Urgente", descricao="vermelho", criado_por=self.creator, prioridade=1
        )
        media = PendenciaTI.objects.create(
            titulo="Media", descricao="amarelo", criado_por=self.creator, prioridade=3
        )
        # Vermelho no topo, verde depois, e as sem cor por ultimo.
        ordenadas = list(PendenciaTI.objects.all())
        self.assertEqual(ordenadas, [urgente, media, baixa, sem_cor])


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class RequisicaoDeleteFilesTests(TestCase):
    """Ao excluir uma requisicao, os arquivos fisicos (fotos e documentos dos
    orcamentos e suborcamentos) devem ser removidos do disco pelos signals."""

    def setUp(self):
        User = get_user_model()
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def _img(self, nome):
        # PNG 1x1 minimo valido para o ImageField aceitar.
        conteudo = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
            b"\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01"
            b"\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        return SimpleUploadedFile(nome, conteudo, content_type="image/png")

    def test_delete_removes_physical_files(self):
        requisicao = RequisicaoContrato.objects.create(titulo="Notebooks", criado_por=self.ti)
        orcamento = OrcamentoContrato.objects.create(
            requisicao=requisicao, titulo="Loja A", foto_produto=self._img("orc.png")
        )
        orc_doc = OrcamentoDocumento.objects.create(
            orcamento=orcamento,
            arquivo=SimpleUploadedFile("orc.pdf", b"pdf", content_type="application/pdf"),
            nome_original="orc.pdf",
        )
        suborcamento = SuborcamentoContrato.objects.create(
            orcamento_pai=orcamento, titulo="Complemento", foto_produto=self._img("sub.png")
        )
        sub_doc = SuborcamentoDocumento.objects.create(
            suborcamento=suborcamento,
            arquivo=SimpleUploadedFile("sub.pdf", b"pdf", content_type="application/pdf"),
            nome_original="sub.pdf",
        )

        caminhos = [
            orcamento.foto_produto.path,
            orc_doc.arquivo.path,
            suborcamento.foto_produto.path,
            sub_doc.arquivo.path,
        ]
        for caminho in caminhos:
            self.assertTrue(os.path.exists(caminho), f"arquivo deveria existir: {caminho}")

        self.client.force_login(self.ti)
        resp = self.client.post(reverse("requisicao_delete", args=[requisicao.id]))
        self.assertEqual(resp.status_code, 200)

        self.assertFalse(RequisicaoContrato.objects.filter(id=requisicao.id).exists())
        for caminho in caminhos:
            self.assertFalse(os.path.exists(caminho), f"arquivo orfao nao removido: {caminho}")


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class OrcamentoAprovacaoTests(TestCase):
    """Aprovacao exclusiva de orcamento dentro de uma requisicao."""

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

        self.requisicao = RequisicaoContrato.objects.create(
            titulo="Notebooks", criado_por=self.ti, status=RequisicaoContrato.STATUS_EM_COTACAO
        )
        self.orc_a = OrcamentoContrato.objects.create(requisicao=self.requisicao, titulo="Loja A")
        self.orc_b = OrcamentoContrato.objects.create(requisicao=self.requisicao, titulo="Loja B")

    def _aprovar(self, orcamento):
        return self.client.post(reverse("orcamento_aprovar", args=[orcamento.id]))

    def test_aprovar_e_exclusivo_e_aguarda_entrega(self):
        self.client.force_login(self.ti)
        resp = self._aprovar(self.orc_a)
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()["aprovado"])

        self.orc_a.refresh_from_db()
        self.orc_b.refresh_from_db()
        self.requisicao.refresh_from_db()
        self.assertTrue(self.orc_a.aprovado)
        self.assertEqual(self.orc_a.aprovado_por, self.ti)
        self.assertFalse(self.orc_b.aprovado)
        # Aprovar move a requisicao para "Aguardando entrega" (nao "Finalizada").
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_AGUARDANDO_ENTREGA)
        # Historico registra a aprovacao.
        self.assertTrue(
            self.requisicao.eventos.filter(tipo=RequisicaoContratoEvento.TIPO_APROVACAO).exists()
        )

        # Aprovar o B remove a aprovacao do A (exclusiva).
        self._aprovar(self.orc_b)
        self.orc_a.refresh_from_db()
        self.orc_b.refresh_from_db()
        self.assertFalse(self.orc_a.aprovado)
        self.assertTrue(self.orc_b.aprovado)

    def test_remover_aprovacao_volta_para_esperando(self):
        self.client.force_login(self.ti)
        self._aprovar(self.orc_a)  # aprova
        resp = self._aprovar(self.orc_a)  # alterna: desaprova
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(resp.json()["aprovado"])

        self.orc_a.refresh_from_db()
        self.requisicao.refresh_from_db()
        self.assertFalse(self.orc_a.aprovado)
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_EM_COTACAO)

    def test_marcar_entregue(self):
        self.client.force_login(self.ti)
        self._aprovar(self.orc_a)  # precisa estar aprovada primeiro
        resp = self.client.post(reverse("requisicao_marcar_entregue", args=[self.requisicao.id]))
        self.assertEqual(resp.status_code, 200)

        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_ENTREGUE)
        self.assertIsNotNone(self.requisicao.entregue_em)
        self.assertEqual(self.requisicao.entregue_por, self.ti)
        self.assertTrue(
            self.requisicao.eventos.filter(tipo=RequisicaoContratoEvento.TIPO_ENTREGA).exists()
        )

        # Nao pode marcar entregue de novo, nem alterar aprovacao depois de entregue.
        self.assertEqual(
            self.client.post(reverse("requisicao_marcar_entregue", args=[self.requisicao.id])).status_code,
            409,
        )
        self.assertEqual(self._aprovar(self.orc_a).status_code, 409)

    def test_marcar_entregue_exige_orcamento_aprovado(self):
        self.client.force_login(self.ti)
        resp = self.client.post(reverse("requisicao_marcar_entregue", args=[self.requisicao.id]))
        self.assertEqual(resp.status_code, 409)  # nenhum orcamento aprovado

    def test_codigo_gerado_continua_do_sistema_antigo(self):
        # Sem requisicoes anteriores no setUp com codigo, a primeira criada aqui
        # deve seguir a numeracao a partir de REQ-00049.
        RequisicaoContrato.objects.all().delete()
        req = RequisicaoContrato.objects.create(titulo="Primeira", criado_por=self.ti)
        self.assertEqual(req.codigo, "REQ-00049")
        req2 = RequisicaoContrato.objects.create(titulo="Segunda", criado_por=self.ti)
        self.assertEqual(req2.codigo, "REQ-00050")

    def test_common_user_cannot_approve(self):
        self.client.force_login(self.common)
        resp = self._aprovar(self.orc_a)
        self.assertEqual(resp.status_code, 403)
        self.orc_a.refresh_from_db()
        self.assertFalse(self.orc_a.aprovado)

    def test_aprovar_requires_post(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("orcamento_aprovar", args=[self.orc_a.id]))
        self.assertEqual(resp.status_code, 405)

    def test_desaprovar_requisicao_limpa_todos_e_volta_para_esperando(self):
        self.client.force_login(self.ti)
        self._aprovar(self.orc_a)  # aprova A -> requisicao "Aguardando entrega"
        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_AGUARDANDO_ENTREGA)

        resp = self.client.post(reverse("requisicao_desaprovar", args=[self.requisicao.id]))
        self.assertEqual(resp.status_code, 200)

        self.orc_a.refresh_from_db()
        self.orc_b.refresh_from_db()
        self.requisicao.refresh_from_db()
        self.assertFalse(self.orc_a.aprovado)
        self.assertIsNone(self.orc_a.aprovado_em)
        self.assertIsNone(self.orc_a.aprovado_por)
        self.assertFalse(self.orc_b.aprovado)
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_EM_COTACAO)
        self.assertTrue(
            self.requisicao.eventos.filter(tipo=RequisicaoContratoEvento.TIPO_APROVACAO).exists()
        )

    def test_desaprovar_sem_orcamento_aprovado_retorna_409(self):
        self.client.force_login(self.ti)
        resp = self.client.post(reverse("requisicao_desaprovar", args=[self.requisicao.id]))
        self.assertEqual(resp.status_code, 409)

    def test_desaprovar_bloqueado_apos_entregue(self):
        self.client.force_login(self.ti)
        self._aprovar(self.orc_a)
        self.client.post(reverse("requisicao_marcar_entregue", args=[self.requisicao.id]))
        resp = self.client.post(reverse("requisicao_desaprovar", args=[self.requisicao.id]))
        self.assertEqual(resp.status_code, 409)
        self.orc_a.refresh_from_db()
        self.assertTrue(self.orc_a.aprovado)  # continua aprovado

    def test_common_user_cannot_desaprovar(self):
        self.client.force_login(self.ti)
        self._aprovar(self.orc_a)
        self.client.logout()
        self.client.force_login(self.common)
        resp = self.client.post(reverse("requisicao_desaprovar", args=[self.requisicao.id]))
        self.assertEqual(resp.status_code, 403)

    def _nao_aprovar(self):
        return self.client.post(reverse("requisicao_nao_aprovar", args=[self.requisicao.id]))

    def test_nao_aprovar_marca_requisicao_como_nao_aprovada(self):
        self.client.force_login(self.ti)
        resp = self._nao_aprovar()
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()["nao_aprovada"])
        self.assertEqual(resp.json()["requisicao_status_label"], "Nao aprovada")

        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_NAO_APROVADA)
        self.assertTrue(
            self.requisicao.eventos.filter(descricao__icontains="NAO APROVADA").exists()
        )

    def test_nao_aprovar_remove_aprovacao_dos_orcamentos(self):
        # Recusar a compra desfaz qualquer aprovacao: nada sera comprado.
        self.client.force_login(self.ti)
        self._aprovar(self.orc_a)
        self._nao_aprovar()

        self.orc_a.refresh_from_db()
        self.requisicao.refresh_from_db()
        self.assertFalse(self.orc_a.aprovado)
        self.assertIsNone(self.orc_a.aprovado_em)
        self.assertIsNone(self.orc_a.aprovado_por)
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_NAO_APROVADA)

    def test_nao_aprovar_de_novo_reabre_a_requisicao(self):
        self.client.force_login(self.ti)
        self._nao_aprovar()
        resp = self._nao_aprovar()  # alterna: reabre
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(resp.json()["nao_aprovada"])

        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_EM_COTACAO)
        self.assertTrue(self.requisicao.eventos.filter(descricao__icontains="reaberta").exists())

    def test_nao_aprovar_bloqueado_apos_entregue(self):
        self.client.force_login(self.ti)
        self._aprovar(self.orc_a)
        self.client.post(reverse("requisicao_marcar_entregue", args=[self.requisicao.id]))
        resp = self._nao_aprovar()
        self.assertEqual(resp.status_code, 409)

        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_ENTREGUE)

    def test_nao_aprovar_exige_post_e_permissao(self):
        self.client.force_login(self.ti)
        self.assertEqual(
            self.client.get(reverse("requisicao_nao_aprovar", args=[self.requisicao.id])).status_code,
            405,
        )
        self.client.logout()
        self.client.force_login(self.common)
        self.assertEqual(self._nao_aprovar().status_code, 403)
        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.status, RequisicaoContrato.STATUS_EM_COTACAO)

    def test_desaprovar_requires_post(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("requisicao_desaprovar", args=[self.requisicao.id]))
        self.assertEqual(resp.status_code, 405)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class RequisicaoEdicaoTests(TestCase):
    """Edicao de requisicao e de seus orcamentos/suborcamentos."""

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

        self.requisicao = RequisicaoContrato.objects.create(
            titulo="Notebooks", tipo=RequisicaoContrato.TIPO_FISICA, criado_por=self.ti
        )
        self.orcamento = OrcamentoContrato.objects.create(
            requisicao=self.requisicao, titulo="Loja A", valor=Decimal("100.00"), quantidade=1
        )
        self.suborcamento = SuborcamentoContrato.objects.create(
            orcamento_pai=self.orcamento, titulo="Complemento", valor=Decimal("10.00"), quantidade=1
        )

    # ----- criacao de suborcamento (replicar em todos) -----
    def test_suborcamento_criado_so_no_orcamento_atual(self):
        # Sem marcar a opcao, o suborcamento entra so no orcamento informado.
        outro = OrcamentoContrato.objects.create(
            requisicao=self.requisicao, titulo="Loja B", valor=Decimal("50.00"), quantidade=1
        )
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("suborcamento_create", args=[self.orcamento.id]),
            {"titulo": "Cabo HDMI", "moeda": "BRL", "valor": "20", "quantidade": "1"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(self.orcamento.suborcamentos.filter(titulo="Cabo HDMI").count(), 1)
        self.assertEqual(outro.suborcamentos.filter(titulo="Cabo HDMI").count(), 0)

    def test_suborcamento_replicado_em_todos_os_orcamentos(self):
        # Marcando a opcao, o mesmo suborcamento e criado em todos os orcamentos
        # principais da requisicao (inclusive o atual).
        outro = OrcamentoContrato.objects.create(
            requisicao=self.requisicao, titulo="Loja B", valor=Decimal("50.00"), quantidade=1
        )
        mais = OrcamentoContrato.objects.create(
            requisicao=self.requisicao, titulo="Loja C", valor=Decimal("70.00"), quantidade=1
        )
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("suborcamento_create", args=[self.orcamento.id]),
            {
                "titulo": "Mouse", "moeda": "BRL", "valor": "30", "quantidade": "1",
                "aplicar_todos_orcamentos": "1",
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["criados"], 3)
        for orc in (self.orcamento, outro, mais):
            self.assertEqual(orc.suborcamentos.filter(titulo="Mouse").count(), 1)

    def test_replicar_suborcamento_exige_ti(self):
        self.client.force_login(self.common)
        resp = self.client.post(
            reverse("suborcamento_create", args=[self.orcamento.id]),
            {"titulo": "X", "moeda": "BRL", "valor": "1", "quantidade": "1",
             "aplicar_todos_orcamentos": "1"},
        )
        self.assertEqual(resp.status_code, 403)

    # ----- requisicao -----
    def test_ti_edita_requisicao(self):
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("requisicao_edit", args=[self.requisicao.id]),
            data=json.dumps({"titulo": "Notebooks Dell", "tipo": "digital", "texto": "Atualizado"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.titulo, "Notebooks Dell")
        self.assertEqual(self.requisicao.tipo, RequisicaoContrato.TIPO_DIGITAL)
        self.assertEqual(self.requisicao.texto, "Atualizado")
        self.assertTrue(
            self.requisicao.eventos.filter(tipo=RequisicaoContratoEvento.TIPO_EDICAO).exists()
        )

    def test_edita_requisicao_titulo_curto(self):
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("requisicao_edit", args=[self.requisicao.id]),
            data=json.dumps({"titulo": "N", "tipo": "fisica", "texto": ""}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)
        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.titulo, "Notebooks")

    def test_common_nao_edita_requisicao(self):
        self.client.force_login(self.common)
        resp = self.client.post(
            reverse("requisicao_edit", args=[self.requisicao.id]),
            data=json.dumps({"titulo": "Hack", "tipo": "fisica", "texto": ""}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_edita_requisicao_requires_post(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("requisicao_edit", args=[self.requisicao.id]))
        self.assertEqual(resp.status_code, 405)

    # ----- orcamento -----
    def test_ti_edita_orcamento(self):
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("orcamento_edit", args=[self.orcamento.id]),
            data={
                "titulo": "Loja B",
                "loja": "Kabum",
                "moeda": "BRL",
                "valor": "250,50",
                "quantidade": "3",
                "frete": "10",
                "desconto": "5",
                "link": "",
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.orcamento.refresh_from_db()
        self.assertEqual(self.orcamento.titulo, "Loja B")
        self.assertEqual(self.orcamento.loja, "Kabum")
        self.assertEqual(self.orcamento.valor, Decimal("250.50"))
        self.assertEqual(self.orcamento.quantidade, 3)

    def test_edita_orcamento_bloqueado_apos_entregue(self):
        self.requisicao.status = RequisicaoContrato.STATUS_ENTREGUE
        self.requisicao.save(update_fields=["status"])
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("orcamento_edit", args=[self.orcamento.id]),
            data={"titulo": "Nova", "moeda": "BRL", "valor": "1", "quantidade": "1"},
        )
        self.assertEqual(resp.status_code, 409)
        self.orcamento.refresh_from_db()
        self.assertEqual(self.orcamento.titulo, "Loja A")

    def test_common_nao_edita_orcamento(self):
        self.client.force_login(self.common)
        resp = self.client.post(
            reverse("orcamento_edit", args=[self.orcamento.id]),
            data={"titulo": "Nova", "moeda": "BRL", "valor": "1", "quantidade": "1"},
        )
        self.assertEqual(resp.status_code, 403)

    def test_edita_orcamento_remove_documento(self):
        doc = OrcamentoDocumento.objects.create(
            orcamento=self.orcamento,
            arquivo=SimpleUploadedFile("orc.pdf", b"pdf", content_type="application/pdf"),
            nome_original="orc.pdf",
        )
        caminho = doc.arquivo.path
        self.assertTrue(os.path.exists(caminho))
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("orcamento_edit", args=[self.orcamento.id]),
            data={
                "titulo": "Loja A",
                "moeda": "BRL",
                "valor": "100",
                "quantidade": "1",
                "remover_documentos": str(doc.id),
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(OrcamentoDocumento.objects.filter(id=doc.id).exists())
        self.assertFalse(os.path.exists(caminho))

    # ----- suborcamento -----
    def test_ti_edita_suborcamento(self):
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("suborcamento_edit", args=[self.suborcamento.id]),
            data={"titulo": "Complemento X", "moeda": "BRL", "valor": "20", "quantidade": "2"},
        )
        self.assertEqual(resp.status_code, 200)
        self.suborcamento.refresh_from_db()
        self.assertEqual(self.suborcamento.titulo, "Complemento X")
        self.assertEqual(self.suborcamento.valor, Decimal("20.00"))
        self.assertEqual(self.suborcamento.quantidade, 2)

    def test_edita_suborcamento_bloqueado_apos_entregue(self):
        self.requisicao.status = RequisicaoContrato.STATUS_ENTREGUE
        self.requisicao.save(update_fields=["status"])
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("suborcamento_edit", args=[self.suborcamento.id]),
            data={"titulo": "Nova", "moeda": "BRL", "valor": "1", "quantidade": "1"},
        )
        self.assertEqual(resp.status_code, 409)


class RequisicaoRenumeracaoTests(TestCase):
    """Migration 0042: renumera os codigos por ordem de criacao (mais recente = 51)."""

    def _renumerar(self):
        import importlib

        from django.apps import apps as global_apps

        mod = importlib.import_module("core.migrations.0042_renumera_codigos_requisicoes")
        mod.renumerar_codigos(global_apps, None)

    def test_renumera_por_ordem_de_criacao(self):
        User = get_user_model()
        ti = User.objects.create_user(username="ti", password="x")

        # Cria 3 requisicoes e força datas de criacao crescentes e codigos "errados".
        base = timezone.now()
        antiga = RequisicaoContrato.objects.create(titulo="Antiga", criado_por=ti)
        meio = RequisicaoContrato.objects.create(titulo="Meio", criado_por=ti)
        recente = RequisicaoContrato.objects.create(titulo="Recente", criado_por=ti)
        RequisicaoContrato.objects.filter(pk=antiga.pk).update(
            criado_em=base - timezone.timedelta(days=2), codigo="REQ-00099"
        )
        RequisicaoContrato.objects.filter(pk=meio.pk).update(
            criado_em=base - timezone.timedelta(days=1), codigo="REQ-00098"
        )
        RequisicaoContrato.objects.filter(pk=recente.pk).update(
            criado_em=base, codigo="REQ-00097"
        )

        self._renumerar()

        antiga.refresh_from_db()
        meio.refresh_from_db()
        recente.refresh_from_db()
        self.assertEqual(antiga.codigo, "REQ-00049")
        self.assertEqual(meio.codigo, "REQ-00050")
        self.assertEqual(recente.codigo, "REQ-00051")

        # A proxima requisicao continua a partir de REQ-00052.
        proxima = RequisicaoContrato.objects.create(titulo="Proxima", criado_por=ti)
        self.assertEqual(proxima.codigo, "REQ-00052")

    def test_renumera_sem_requisicoes_nao_falha(self):
        self._renumerar()  # banco vazio: nao deve levantar erro
        self.assertEqual(RequisicaoContrato.objects.count(), 0)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class RequisicaoBuscaTests(TestCase):
    """Busca inteligente da lista de requisicoes (texto pesquisavel no data-search).

    A lista mostra so codigo/titulo/status, mas a pesquisa filtra por qualquer
    coisa da requisicao. O filtro roda no navegador sobre o texto que a view
    monta em `_requisicao_busca_texto`; aqui garantimos que esse texto leva os
    dados da requisicao e dos orcamentos/suborcamentos.
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x", first_name="Fabiano")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

        self.requisicao = RequisicaoContrato.objects.create(
            titulo="Notebooks para o RH",
            tipo=RequisicaoContrato.TIPO_FISICA,
            texto="Substituir as maquinas antigas do setor",
            criado_por=self.ti,
        )
        self.orcamento = OrcamentoContrato.objects.create(
            requisicao=self.requisicao,
            titulo="Dell Vostro",
            loja="Kabum",
            link="https://loja.exemplo/produto-1",
            valor=Decimal("4321.99"),
            quantidade=2,
        )
        SuborcamentoContrato.objects.create(
            orcamento_pai=self.orcamento,
            titulo="Mouse sem fio",
            loja="Mercado Livre",
            valor=Decimal("89.90"),
            quantidade=1,
        )
        OrcamentoDocumento.objects.create(
            orcamento=self.orcamento,
            arquivo=SimpleUploadedFile("proposta.pdf", b"conteudo"),
            nome_original="proposta-dell.pdf",
        )

    def _busca(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("contratos_dashboard"))
        self.assertEqual(resp.status_code, 200)
        linhas = [r for r in resp.context["requisicoes"] if r["id"] == self.requisicao.id]
        self.assertEqual(len(linhas), 1)
        return linhas[0]["busca"]

    def test_busca_leva_dados_da_requisicao(self):
        texto = self._busca()
        self.assertIn(self.requisicao.codigo, texto)
        self.assertIn("Notebooks para o RH", texto)
        self.assertIn("Fisica", texto)
        self.assertIn("Aberta", texto)
        self.assertIn("maquinas antigas", texto)
        self.assertIn("Fabiano", texto)

    def test_busca_leva_orcamentos_suborcamentos_e_documentos(self):
        texto = self._busca()
        # Orcamento: titulo, loja, link, valor cru e total formatado.
        self.assertIn("Dell Vostro", texto)
        self.assertIn("Kabum", texto)
        self.assertIn("https://loja.exemplo/produto-1", texto)
        self.assertIn("4321.99", texto)
        self.assertIn("R$ 8.643,98", texto)  # 4321,99 x 2
        # Suborcamento e o nome do documento anexado.
        self.assertIn("Mouse sem fio", texto)
        self.assertIn("Mercado Livre", texto)
        self.assertIn("proposta-dell.pdf", texto)

    def test_busca_acompanha_aprovacao_e_status(self):
        self.client.force_login(self.ti)
        resp = self.client.post(reverse("orcamento_aprovar", args=[self.orcamento.id]))
        self.assertEqual(resp.status_code, 200)
        texto = self._busca()
        self.assertIn("Aguardando entrega", texto)
        self.assertIn("aprovado", texto)

    def test_campo_de_busca_aparece_na_tela(self):
        self.client.force_login(self.ti)
        html = self.client.get(reverse("contratos_dashboard")).content.decode()
        self.assertIn('id="requisicaoSearch"', html)
        self.assertIn("data-search=", html)

    def test_usuario_comum_nao_acessa_a_lista(self):
        self.client.force_login(self.common)
        resp = self.client.get(reverse("contratos_dashboard"))
        self.assertEqual(resp.status_code, 302)


class ContaEmailImportTests(TestCase):
    """Importacao da lista de contas de e-mail (upsert por e-mail) e permissoes."""

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def _csv(self, deptos):
        header = (
            "First Name [Required],Last Name [Required],Email Address [Required],"
            "Status [READ ONLY],Department,2sv Enrolled [READ ONLY]\n"
        )
        linhas = [
            f"Joao,Silva,joao.silva@x.com,Active,{deptos[0]},True",
            f"Maria,Souza,maria.souza@x.com,Suspended,{deptos[1]},False",
        ]
        conteudo = (header + "\n".join(linhas)).encode("utf-8")
        return SimpleUploadedFile("lista.csv", conteudo, content_type="text/csv")

    def test_ti_import_creates_and_upserts(self):
        self.client.force_login(self.ti)
        resp = self.client.post(reverse("email_import"), {"arquivo": self._csv(["TI", "RH"])})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(ContaEmail.objects.count(), 2)

        joao = ContaEmail.objects.get(email="joao.silva@x.com")
        self.assertEqual(joao.primeiro_nome, "Joao")
        self.assertEqual(joao.departamento, "TI")
        self.assertTrue(joao.dois_fatores_inscrito)
        self.assertTrue(joao.is_ativo)
        self.assertFalse(ContaEmail.objects.get(email="maria.souza@x.com").is_ativo)

        # Reimportar com o mesmo e-mail atualiza (nao duplica).
        self.client.post(reverse("email_import"), {"arquivo": self._csv(["Infra", "RH"])})
        self.assertEqual(ContaEmail.objects.count(), 2)
        joao.refresh_from_db()
        self.assertEqual(joao.departamento, "Infra")

    def test_common_user_cannot_import(self):
        self.client.force_login(self.common)
        resp = self.client.post(reverse("email_import"), {"arquivo": self._csv(["TI", "RH"])})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(ContaEmail.objects.count(), 0)

    def test_common_user_cannot_access_dashboard(self):
        self.client.force_login(self.common)
        resp = self.client.get(reverse("emails_dashboard"))
        self.assertEqual(resp.status_code, 302)  # redirecionado (sem permissao TI)


class RamalCreateTests(TestCase):
    """Cadastro de ramal (e-mail vindo de uma ContaEmail) e permissoes.

    Obs.: o banco de teste ja vem com os ramais do seed (migration 0013), por
    isso os testes comparam a contagem antes/depois em vez de assumir zero.
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))
        self.conta = ContaEmail.objects.create(
            email="novo.contato@x.com", primeiro_nome="Novo", sobrenome="Contato"
        )

    def test_ti_creates_ramal_with_selected_email(self):
        self.client.force_login(self.ti)
        antes = Ramal.objects.count()
        resp = self.client.post(
            reverse("ramal_create"),
            {"colaborador": "Zzz Teste", "setor": "TI", "telefone": "123", "ramal": "9000", "conta_email": self.conta.id},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Ramal.objects.count(), antes + 1)

        ramal = Ramal.objects.get(colaborador="Zzz Teste")
        self.assertEqual(ramal.email, "novo.contato@x.com")  # puxado da conta selecionada
        self.assertEqual(ramal.conta_email, self.conta)
        self.assertEqual(ramal.ramal, "9000")

    def test_create_requires_colaborador(self):
        self.client.force_login(self.ti)
        antes = Ramal.objects.count()
        resp = self.client.post(reverse("ramal_create"), {"colaborador": "", "setor": "TI"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Ramal.objects.count(), antes)  # nada criado

    def test_common_user_cannot_create(self):
        self.client.force_login(self.common)
        antes = Ramal.objects.count()
        resp = self.client.post(
            reverse("ramal_create"), {"colaborador": "Hacker", "setor": "X"}
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Ramal.objects.count(), antes)

    def test_common_user_cannot_access_dashboard(self):
        self.client.force_login(self.common)
        resp = self.client.get(reverse("ramais_dashboard"))
        self.assertEqual(resp.status_code, 302)

    def test_create_with_free_email(self):
        """E-mail digitado livremente (sem conta selecionada) e aceito."""
        self.client.force_login(self.ti)
        self.client.post(
            reverse("ramal_create"),
            {"colaborador": "Sala Teste", "setor": "Reuniao", "email": "livre@x.com"},
        )
        ramal = Ramal.objects.get(colaborador="Sala Teste")
        self.assertEqual(ramal.email, "livre@x.com")
        self.assertIsNone(ramal.conta_email)

    def test_ti_updates_ramal(self):
        self.client.force_login(self.ti)
        ramal = Ramal.objects.create(colaborador="Antigo", setor="X", ramal="1000")
        resp = self.client.post(
            reverse("ramal_update", args=[ramal.id]),
            {"colaborador": "Novo Nome", "setor": "TI", "telefone": "9", "ramal": "1001", "conta_email": self.conta.id},
        )
        self.assertEqual(resp.status_code, 302)
        ramal.refresh_from_db()
        self.assertEqual(ramal.colaborador, "Novo Nome")
        self.assertEqual(ramal.ramal, "1001")
        self.assertEqual(ramal.email, self.conta.email)  # veio da conta selecionada
        self.assertEqual(ramal.conta_email, self.conta)

    def test_ti_deletes_ramal(self):
        self.client.force_login(self.ti)
        ramal = Ramal.objects.create(colaborador="Excluir", setor="X")
        antes = Ramal.objects.count()
        resp = self.client.post(reverse("ramal_delete", args=[ramal.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Ramal.objects.count(), antes - 1)
        self.assertFalse(Ramal.objects.filter(id=ramal.id).exists())

    def test_common_user_cannot_update_or_delete(self):
        ramal = Ramal.objects.create(colaborador="Protegido", setor="X")
        self.client.force_login(self.common)
        self.client.post(reverse("ramal_update", args=[ramal.id]), {"colaborador": "Hack"})
        self.client.post(reverse("ramal_delete", args=[ramal.id]))
        ramal.refresh_from_db()
        self.assertEqual(ramal.colaborador, "Protegido")  # inalterado
        self.assertTrue(Ramal.objects.filter(id=ramal.id).exists())  # nao excluido


class LicencaTests(TestCase):
    """Modulo Licencas: CRUD de software e licenca, prazos e permissoes.

    Obs.: o banco de teste ja vem com os softwares/licencas do seed (migration
    0015), por isso os testes comparam contagem antes/depois.
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def test_ti_creates_software(self):
        self.client.force_login(self.ti)
        antes = LicencaSoftware.objects.count()
        resp = self.client.post(
            reverse("licenca_software_create"),
            {"nome": "Photoshop 2026", "quantidade_licencas": 4, "observacoes": "Assinatura"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(LicencaSoftware.objects.count(), antes + 1)
        soft = LicencaSoftware.objects.get(nome="Photoshop 2026")
        self.assertEqual(soft.quantidade_licencas, 4)
        self.assertEqual(soft.criado_por, self.ti)

    def test_software_create_requires_name(self):
        self.client.force_login(self.ti)
        antes = LicencaSoftware.objects.count()
        resp = self.client.post(reverse("licenca_software_create"), {"nome": "", "quantidade_licencas": 1})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(LicencaSoftware.objects.count(), antes)  # nada criado

    def test_ti_creates_license_and_expiration(self):
        self.client.force_login(self.ti)
        soft = LicencaSoftware.objects.create(nome="CorelDRAW", quantidade_licencas=2, criado_por=self.ti)
        resp = self.client.post(
            reverse("licenca_create"),
            {
                "software": soft.id,
                "usuario_atribuido": "Fulano",
                "serial": "ABC-123",
                "email_vinculado": "fulano@x.com",
                "tipo_expiracao": "expira_em",
                "expira_em": "2027-01-01",
                "forma_pagamento": "Boleto",
                "final_cartao": "9999",
            },
        )
        self.assertEqual(resp.status_code, 302)
        lic = Licenca.objects.get(serial="ABC-123")
        self.assertEqual(lic.software, soft)
        self.assertEqual(lic.expira_label, "01/01/2027")
        self.assertEqual(lic.final_cartao, "9999")

    def test_indeterminado_clears_expira(self):
        """Prazo indeterminado ignora a data enviada."""
        self.client.force_login(self.ti)
        soft = LicencaSoftware.objects.create(nome="Zoom", quantidade_licencas=1, criado_por=self.ti)
        self.client.post(
            reverse("licenca_create"),
            {"software": soft.id, "tipo_expiracao": "indeterminado", "expira_em": "2027-05-05", "usuario_atribuido": "X"},
        )
        lic = Licenca.objects.get(software=soft, usuario_atribuido="X")
        self.assertIsNone(lic.expira_em)
        self.assertEqual(lic.expira_label, "Indeterminado")

    def test_license_requires_valid_software(self):
        self.client.force_login(self.ti)
        antes = Licenca.objects.count()
        resp = self.client.post(reverse("licenca_create"), {"software": 999999, "usuario_atribuido": "X"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Licenca.objects.count(), antes)  # nada criado

    def test_ti_updates_and_deletes_license(self):
        self.client.force_login(self.ti)
        soft = LicencaSoftware.objects.create(nome="Slack", quantidade_licencas=1, criado_por=self.ti)
        lic = Licenca.objects.create(software=soft, usuario_atribuido="Antes", criado_por=self.ti)
        self.client.post(
            reverse("licenca_update", args=[lic.id]),
            {"software": soft.id, "usuario_atribuido": "Depois", "tipo_expiracao": "indeterminado"},
        )
        lic.refresh_from_db()
        self.assertEqual(lic.usuario_atribuido, "Depois")

        resp = self.client.post(reverse("licenca_delete", args=[lic.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Licenca.objects.filter(id=lic.id).exists())

    def test_delete_software_cascades_licenses(self):
        self.client.force_login(self.ti)
        soft = LicencaSoftware.objects.create(nome="Trello", quantidade_licencas=1, criado_por=self.ti)
        lic = Licenca.objects.create(software=soft, usuario_atribuido="Y", criado_por=self.ti)
        resp = self.client.post(reverse("licenca_software_delete", args=[soft.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(LicencaSoftware.objects.filter(id=soft.id).exists())
        self.assertFalse(Licenca.objects.filter(id=lic.id).exists())  # cascata

    def test_common_user_cannot_access_or_change(self):
        soft = LicencaSoftware.objects.create(nome="Protegido", quantidade_licencas=1, criado_por=self.ti)
        self.client.force_login(self.common)
        # Dashboard redireciona
        self.assertEqual(self.client.get(reverse("licencas_dashboard")).status_code, 302)
        # Nao cria software nem licenca
        antes_s = LicencaSoftware.objects.count()
        antes_l = Licenca.objects.count()
        self.client.post(reverse("licenca_software_create"), {"nome": "Hack", "quantidade_licencas": 1})
        self.client.post(reverse("licenca_create"), {"software": soft.id, "usuario_atribuido": "Hack"})
        self.assertEqual(LicencaSoftware.objects.count(), antes_s)
        self.assertEqual(Licenca.objects.count(), antes_l)
        # Nao exclui
        self.client.post(reverse("licenca_software_delete", args=[soft.id]))
        self.assertTrue(LicencaSoftware.objects.filter(id=soft.id).exists())

    def test_dashboard_lists_seeded_software(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("licencas_dashboard"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "AutoCAD 2014 Full")


class EnderecoIPTests(TestCase):
    """Modulo IPs: CRUD, unicidade do IP, categoria e permissoes.

    Obs.: o banco de teste ja vem com os IPs do seed (migration 0017).
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def test_ti_creates_ip(self):
        self.client.force_login(self.ti)
        antes = EnderecoIP.objects.count()
        resp = self.client.post(
            reverse("ip_create"),
            {"categoria": "servers", "endereco_ip": "10.0.0.9", "nome": "SRV-TESTE", "mac": "AA:BB:CC:DD:EE:FF"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(EnderecoIP.objects.count(), antes + 1)
        ip = EnderecoIP.objects.get(endereco_ip="10.0.0.9")
        self.assertEqual(ip.categoria, "servers")
        self.assertEqual(ip.criado_por, self.ti)

    def test_create_requires_ip_and_valid_category(self):
        self.client.force_login(self.ti)
        antes = EnderecoIP.objects.count()
        # sem endereco
        self.client.post(reverse("ip_create"), {"categoria": "servers", "endereco_ip": ""})
        # categoria invalida
        self.client.post(reverse("ip_create"), {"categoria": "xpto", "endereco_ip": "10.0.0.10"})
        self.assertEqual(EnderecoIP.objects.count(), antes)  # nada criado

    def test_duplicate_ip_is_rejected(self):
        self.client.force_login(self.ti)
        EnderecoIP.objects.create(categoria="wifi", endereco_ip="10.0.0.20", criado_por=self.ti)
        antes = EnderecoIP.objects.count()
        resp = self.client.post(
            reverse("ip_create"), {"categoria": "servers", "endereco_ip": "10.0.0.20"}
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(EnderecoIP.objects.count(), antes)  # nao duplicou

    def test_ti_updates_and_deletes_ip(self):
        self.client.force_login(self.ti)
        ip = EnderecoIP.objects.create(categoria="printers", endereco_ip="10.0.0.30", nome="Antes", criado_por=self.ti)
        resp = self.client.post(
            reverse("ip_update", args=[ip.id]),
            {"categoria": "printers", "endereco_ip": "10.0.0.30", "nome": "Depois"},
        )
        self.assertEqual(resp.status_code, 302)
        ip.refresh_from_db()
        self.assertEqual(ip.nome, "Depois")

        resp = self.client.post(reverse("ip_delete", args=[ip.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(EnderecoIP.objects.filter(id=ip.id).exists())

    def test_update_keeps_own_ip_without_duplicate_error(self):
        """Editar mantendo o mesmo IP nao dispara falso positivo de duplicidade."""
        self.client.force_login(self.ti)
        ip = EnderecoIP.objects.create(categoria="switches", endereco_ip="10.0.0.40", criado_por=self.ti)
        resp = self.client.post(
            reverse("ip_update", args=[ip.id]),
            {"categoria": "switches", "endereco_ip": "10.0.0.40", "nome": "Switch Novo"},
        )
        self.assertEqual(resp.status_code, 302)
        ip.refresh_from_db()
        self.assertEqual(ip.nome, "Switch Novo")

    def test_common_user_cannot_access_or_change(self):
        ip = EnderecoIP.objects.create(categoria="wifi", endereco_ip="10.0.0.50", criado_por=self.ti)
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("ips_dashboard")).status_code, 302)
        antes = EnderecoIP.objects.count()
        self.client.post(reverse("ip_create"), {"categoria": "servers", "endereco_ip": "10.0.0.99"})
        self.assertEqual(EnderecoIP.objects.count(), antes)
        self.client.post(reverse("ip_delete", args=[ip.id]))
        self.assertTrue(EnderecoIP.objects.filter(id=ip.id).exists())

    def test_dashboard_renders_with_seeded_data(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("ips_dashboard"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "ips-table")


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class ServicoFeitoTests(TestCase):
    """Modulo Servicos feitos: CRUD, anexos, valor BR e permissoes.

    Obs.: o banco de teste ja vem com os servicos do seed (migration 0019),
    por isso os testes comparam contagem antes/depois.
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def _arquivo(self, nome="nota.pdf"):
        return SimpleUploadedFile(nome, b"%PDF-1.4 conteudo", content_type="application/pdf")

    def test_ti_creates_service_with_attachments(self):
        self.client.force_login(self.ti)
        antes = ServicoFeito.objects.count()
        resp = self.client.post(
            reverse("servico_feito_create"),
            {
                "nome_servico": "Troca de switch",
                "empresa": "Acme",
                "data_servico": "2026-05-10",
                "valor": "1.234,56",
                "descricao": "Substituicao do switch principal",
                "anexos": [self._arquivo("nf1.pdf"), self._arquivo("nf2.pdf")],
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(ServicoFeito.objects.count(), antes + 1)
        servico = ServicoFeito.objects.get(nome_servico="Troca de switch")
        self.assertEqual(str(servico.valor), "1234.56")  # valor BR convertido
        self.assertEqual(servico.anexos.count(), 2)
        self.assertEqual(servico.criado_por, self.ti)

    def test_create_requires_name(self):
        self.client.force_login(self.ti)
        antes = ServicoFeito.objects.count()
        resp = self.client.post(reverse("servico_feito_create"), {"nome_servico": "", "valor": "10"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(ServicoFeito.objects.count(), antes)

    def test_valor_display_pt_br(self):
        s = ServicoFeito.objects.create(nome_servico="X", valor="20796.01", criado_por=self.ti)
        self.assertEqual(s.valor_display, "20.796,01")

    def test_ti_updates_and_adds_attachment(self):
        self.client.force_login(self.ti)
        s = ServicoFeito.objects.create(nome_servico="Antes", valor="100", criado_por=self.ti)
        resp = self.client.post(
            reverse("servico_feito_update", args=[s.id]),
            {"nome_servico": "Depois", "valor": "200", "data_servico": "2026-06-01", "anexos": [self._arquivo()]},
        )
        self.assertEqual(resp.status_code, 302)
        s.refresh_from_db()
        self.assertEqual(s.nome_servico, "Depois")
        self.assertEqual(str(s.valor), "200.00")
        self.assertEqual(s.anexos.count(), 1)

    def test_delete_service_cascades_attachments(self):
        self.client.force_login(self.ti)
        s = ServicoFeito.objects.create(nome_servico="Apagar", valor="1", criado_por=self.ti)
        a = ServicoFeitoAnexo.objects.create(servico=s, arquivo=self._arquivo(), nome_original="x.pdf")
        resp = self.client.post(reverse("servico_feito_delete", args=[s.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(ServicoFeito.objects.filter(id=s.id).exists())
        self.assertFalse(ServicoFeitoAnexo.objects.filter(id=a.id).exists())

    def test_delete_single_attachment(self):
        self.client.force_login(self.ti)
        s = ServicoFeito.objects.create(nome_servico="Com anexo", valor="1", criado_por=self.ti)
        a = ServicoFeitoAnexo.objects.create(servico=s, arquivo=self._arquivo(), nome_original="x.pdf")
        resp = self.client.post(reverse("servico_feito_anexo_delete", args=[a.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(ServicoFeitoAnexo.objects.filter(id=a.id).exists())
        self.assertTrue(ServicoFeito.objects.filter(id=s.id).exists())  # servico permanece

    def test_detail_json_and_download(self):
        self.client.force_login(self.ti)
        s = ServicoFeito.objects.create(nome_servico="Detalhe", empresa="Y", valor="50", criado_por=self.ti)
        a = ServicoFeitoAnexo.objects.create(servico=s, arquivo=self._arquivo(), nome_original="doc.pdf")
        resp = self.client.get(reverse("servico_feito_detail", args=[s.id]))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data["ok"])
        self.assertEqual(data["valor_display"], "50,00")
        self.assertEqual(len(data["anexos"]), 1)
        # Download protegido funciona para TI
        dl = self.client.get(reverse("servico_feito_anexo_download", args=[a.id]))
        self.assertEqual(dl.status_code, 200)

    def test_common_user_blocked(self):
        s = ServicoFeito.objects.create(nome_servico="Protegido", valor="1", criado_por=self.ti)
        a = ServicoFeitoAnexo.objects.create(servico=s, arquivo=self._arquivo(), nome_original="x.pdf")
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("servicos_feitos_dashboard")).status_code, 302)
        self.assertEqual(self.client.get(reverse("servico_feito_detail", args=[s.id])).status_code, 403)
        self.assertEqual(self.client.get(reverse("servico_feito_anexo_download", args=[a.id])).status_code, 404)
        antes = ServicoFeito.objects.count()
        self.client.post(reverse("servico_feito_create"), {"nome_servico": "Hack", "valor": "1"})
        self.assertEqual(ServicoFeito.objects.count(), antes)
        self.client.post(reverse("servico_feito_delete", args=[s.id]))
        self.assertTrue(ServicoFeito.objects.filter(id=s.id).exists())


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class ContratoTests(TestCase):
    """Modulo Contratos: CRUD, anexos, valor BR, periodicidade e permissoes.

    Obs.: o banco de teste ja vem com os contratos do seed (migration 0021).
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def _arquivo(self, nome="contrato.pdf"):
        return SimpleUploadedFile(nome, b"%PDF-1.4 x", content_type="application/pdf")

    def test_ti_creates_contract_with_attachments(self):
        self.client.force_login(self.ti)
        antes = Contrato.objects.count()
        resp = self.client.post(
            reverse("contrato_ti_create"),
            {
                "nome": "Contrato Teste Unico",
                "valor": "4.798,03",
                "periodicidade": "mensal",
                "forma_pagamento": "Boleto",
                "inicio": "2026-02-28",
                "fim": "2029-02-28",
                "anexos": [self._arquivo("nf.pdf")],
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Contrato.objects.count(), antes + 1)
        ctr = Contrato.objects.get(nome="Contrato Teste Unico")
        self.assertEqual(str(ctr.valor), "4798.03")
        self.assertEqual(ctr.periodicidade, "mensal")
        self.assertEqual(ctr.anexos.count(), 1)
        self.assertTrue(ctr.esta_ativo)
        self.assertEqual(ctr.criado_por, self.ti)

    def test_create_requires_name(self):
        self.client.force_login(self.ti)
        antes = Contrato.objects.count()
        self.client.post(reverse("contrato_ti_create"), {"nome": "", "valor": "10"})
        self.assertEqual(Contrato.objects.count(), antes)

    def test_valor_display_and_optional_value(self):
        c1 = Contrato.objects.create(nome="Com valor", valor="1716.20", criado_por=self.ti)
        self.assertEqual(c1.valor_display, "1.716,20")
        c2 = Contrato.objects.create(nome="Sem valor", criado_por=self.ti)
        self.assertEqual(c2.valor_display, "-")

    def test_encerrado_marks_inactive(self):
        self.client.force_login(self.ti)
        c = Contrato.objects.create(nome="Ativo", valor="1", criado_por=self.ti)
        self.assertTrue(c.esta_ativo)
        self.client.post(
            reverse("contrato_ti_update", args=[c.id]),
            {"nome": "Ativo", "valor": "1", "periodicidade": "mensal", "encerrado_em": "2026-05-26"},
        )
        c.refresh_from_db()
        self.assertFalse(c.esta_ativo)

    def test_delete_cascades_attachments(self):
        self.client.force_login(self.ti)
        c = Contrato.objects.create(nome="Apagar", valor="1", criado_por=self.ti)
        a = ContratoAnexo.objects.create(contrato=c, arquivo=self._arquivo(), nome_original="x.pdf")
        resp = self.client.post(reverse("contrato_ti_delete", args=[c.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Contrato.objects.filter(id=c.id).exists())
        self.assertFalse(ContratoAnexo.objects.filter(id=a.id).exists())

    def test_delete_single_attachment(self):
        self.client.force_login(self.ti)
        c = Contrato.objects.create(nome="Com anexo", valor="1", criado_por=self.ti)
        a = ContratoAnexo.objects.create(contrato=c, arquivo=self._arquivo(), nome_original="x.pdf")
        resp = self.client.post(reverse("contrato_ti_anexo_delete", args=[a.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(ContratoAnexo.objects.filter(id=a.id).exists())
        self.assertTrue(Contrato.objects.filter(id=c.id).exists())

    def test_detail_json_and_download(self):
        self.client.force_login(self.ti)
        c = Contrato.objects.create(nome="Detalhe", valor="50", periodicidade="anual", criado_por=self.ti)
        a = ContratoAnexo.objects.create(contrato=c, arquivo=self._arquivo(), nome_original="doc.pdf")
        resp = self.client.get(reverse("contrato_ti_detail", args=[c.id]))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data["ok"])
        self.assertEqual(data["valor_display"], "50,00")
        self.assertEqual(data["periodicidade"], "Anual")
        self.assertEqual(len(data["anexos"]), 1)
        dl = self.client.get(reverse("contrato_ti_anexo_download", args=[a.id]))
        self.assertEqual(dl.status_code, 200)

    def test_common_user_blocked(self):
        c = Contrato.objects.create(nome="Protegido", valor="1", criado_por=self.ti)
        a = ContratoAnexo.objects.create(contrato=c, arquivo=self._arquivo(), nome_original="x.pdf")
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("contratos_ti_dashboard")).status_code, 302)
        self.assertEqual(self.client.get(reverse("contrato_ti_detail", args=[c.id])).status_code, 403)
        self.assertEqual(self.client.get(reverse("contrato_ti_anexo_download", args=[a.id])).status_code, 404)
        antes = Contrato.objects.count()
        self.client.post(reverse("contrato_ti_create"), {"nome": "Hack", "valor": "1"})
        self.assertEqual(Contrato.objects.count(), antes)
        self.client.post(reverse("contrato_ti_delete", args=[c.id]))
        self.assertTrue(Contrato.objects.filter(id=c.id).exists())


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class FuturaDigitalTests(TestCase):
    """Modulo Futura Digital: regra de cobranca, CRUD e permissoes.

    Obs.: o banco de teste ja vem com as faturas do seed (migration 0023).
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def _criar(self, mes, total, cor, **extra):
        dados = {
            "mes_referencia": mes,
            "copias_total": str(total),
            "copias_cor": str(cor),
            "franquia_copias": "23000",
            "franquia_valor": "1610,00",
            "valor_copia_excedente": "0,07",
            "valor_copia_cor": "0,75",
        }
        dados.update(extra)
        return self.client.post(reverse("futura_digital_create"), dados)

    def test_billing_rule(self):
        """valor = franquia + excedentes*rate_exc + cor*rate_cor;
        excedentes = producao total - franquia (as coloridas NAO saem da base:
        elas contam no volume da franquia e ainda pagam a taxa de cor)."""
        self.client.force_login(self.ti)
        # Numeros da memoria de calculo/NF: producao 56.141, coloridas 498.
        resp = self._criar("2026-07", 56141, 498)
        self.assertEqual(resp.status_code, 302)

        f = FuturaDigital.objects.get(mes_referencia="2026-07-01")
        # 56141 - 23000 = 33141 (as 498 coloridas continuam na conta)
        self.assertEqual(f.copias_excedentes, 33141)
        # 1610 + 33141*0.07 + 498*0.75 = 4303.37 (valor da NF)
        self.assertEqual(str(f.valor_pago), "4303.37")
        self.assertEqual(f.criado_por, self.ti)
        # A producao P&B fica so como informacao, fora do calculo.
        self.assertEqual(f.copias_pb, 55643)

    def test_billing_rule_bate_com_a_fatura_de_maio(self):
        # Fatura 05/2026 (producao 44.436, coloridas 204) = R$ 3.263,52.
        self.client.force_login(self.ti)
        self._criar("2026-08", 44436, 204)
        f = FuturaDigital.objects.get(mes_referencia="2026-08-01")
        self.assertEqual(f.copias_excedentes, 21436)
        self.assertEqual(str(f.valor_pago), "3263.52")

    def test_producao_abaixo_da_franquia_nao_gera_excedente(self):
        self.client.force_login(self.ti)
        self._criar("2026-09", 20000, 300)
        f = FuturaDigital.objects.get(mes_referencia="2026-09-01")
        self.assertEqual(f.copias_excedentes, 0)
        # So franquia + coloridas: 1610 + 300*0.75 = 1835.00
        self.assertEqual(str(f.valor_pago), "1835.00")

    def test_month_normalized_to_first_day(self):
        self.client.force_login(self.ti)
        self.client.post(
            reverse("futura_digital_create"),
            {"mes_referencia": "2026-08", "copias_total": "1000", "copias_cor": "0"},
        )
        f = FuturaDigital.objects.get(mes_referencia="2026-08-01")
        self.assertEqual(f.mes_referencia.day, 1)

    def test_no_excess_when_below_franchise(self):
        self.client.force_login(self.ti)
        self.client.post(
            reverse("futura_digital_create"),
            {"mes_referencia": "2026-09", "copias_total": "10000", "copias_cor": "0", "franquia_copias": "23000"},
        )
        f = FuturaDigital.objects.get(mes_referencia="2026-09-01")
        self.assertEqual(f.copias_excedentes, 0)
        self.assertEqual(str(f.valor_pago), "1610.00")  # so a franquia

    def test_create_requires_month(self):
        self.client.force_login(self.ti)
        antes = FuturaDigital.objects.count()
        self.client.post(reverse("futura_digital_create"), {"mes_referencia": "", "copias_total": "10"})
        self.assertEqual(FuturaDigital.objects.count(), antes)

    def test_update_recalculates(self):
        self.client.force_login(self.ti)
        f = FuturaDigital.objects.create(mes_referencia="2027-01-01", copias_total=1000, copias_cor=0)
        f.recalcular(); f.save()
        self.client.post(
            reverse("futura_digital_update", args=[f.id]),
            {"mes_referencia": "2027-01", "copias_total": "60000", "copias_cor": "500", "franquia_copias": "23000"},
        )
        f.refresh_from_db()
        # 60000 - 23000 = 37000 (as 500 coloridas continuam na base do excedente)
        self.assertEqual(f.copias_excedentes, 37000)

    def test_delete(self):
        self.client.force_login(self.ti)
        f = FuturaDigital.objects.create(mes_referencia="2027-02-01", copias_total=1000)
        resp = self.client.post(reverse("futura_digital_delete", args=[f.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(FuturaDigital.objects.filter(id=f.id).exists())

    def test_common_user_blocked(self):
        f = FuturaDigital.objects.create(mes_referencia="2027-03-01", copias_total=1000)
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("futura_digital_dashboard")).status_code, 302)
        antes = FuturaDigital.objects.count()
        self.client.post(reverse("futura_digital_create"), {"mes_referencia": "2027-04", "copias_total": "1"})
        self.assertEqual(FuturaDigital.objects.count(), antes)
        self.client.post(reverse("futura_digital_delete", args=[f.id]))
        self.assertTrue(FuturaDigital.objects.filter(id=f.id).exists())

    def test_dashboard_renders_chart_data(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("futura_digital_dashboard"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "fdSerie")
        self.assertContains(resp, "fd-chart")


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class DicaTests(TestCase):
    """Modulo Dicas: CRUD, categorias, anexo e permissoes.

    Obs.: o banco de teste ja vem com as dicas do seed (migration 0025).
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def _img(self, nome="print.png"):
        return SimpleUploadedFile(nome, b"\x89PNG\r\n\x1a\n fake", content_type="image/png")

    def test_ti_creates_dica(self):
        self.client.force_login(self.ti)
        antes = Dica.objects.count()
        resp = self.client.post(
            reverse("dica_create"),
            {"categoria": "resolucao", "titulo": "Reset do servidor X", "conteudo": "Passo 1\nPasso 2", "anexo": self._img()},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Dica.objects.count(), antes + 1)
        d = Dica.objects.get(titulo="Reset do servidor X")
        self.assertEqual(d.categoria, "resolucao")
        self.assertTrue(d.anexo)
        self.assertEqual(d.criado_por, self.ti)

    def test_create_requires_title(self):
        self.client.force_login(self.ti)
        antes = Dica.objects.count()
        self.client.post(reverse("dica_create"), {"categoria": "geral", "titulo": ""})
        self.assertEqual(Dica.objects.count(), antes)

    def test_invalid_category_falls_back_to_geral(self):
        self.client.force_login(self.ti)
        self.client.post(reverse("dica_create"), {"categoria": "xpto", "titulo": "Dica Z"})
        self.assertEqual(Dica.objects.get(titulo="Dica Z").categoria, "geral")

    def test_update_and_remove_attachment(self):
        self.client.force_login(self.ti)
        d = Dica.objects.create(categoria="geral", titulo="Antes", conteudo="x", anexo=self._img())
        self.assertTrue(d.anexo)
        self.client.post(
            reverse("dica_update", args=[d.id]),
            {"categoria": "configuracao", "titulo": "Depois", "conteudo": "y", "remover_anexo": "1"},
        )
        d.refresh_from_db()
        self.assertEqual(d.titulo, "Depois")
        self.assertEqual(d.categoria, "configuracao")
        self.assertFalse(d.anexo)

    def test_delete(self):
        self.client.force_login(self.ti)
        d = Dica.objects.create(categoria="geral", titulo="Apagar")
        resp = self.client.post(reverse("dica_delete", args=[d.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Dica.objects.filter(id=d.id).exists())

    def test_anexo_download_protected(self):
        d = Dica.objects.create(categoria="geral", titulo="Com anexo", anexo=self._img())
        self.client.force_login(self.ti)
        self.assertEqual(self.client.get(reverse("dica_anexo", args=[d.id])).status_code, 200)
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("dica_anexo", args=[d.id])).status_code, 404)

    def test_common_user_blocked(self):
        d = Dica.objects.create(categoria="geral", titulo="Protegida")
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("dicas_dashboard")).status_code, 302)
        antes = Dica.objects.count()
        self.client.post(reverse("dica_create"), {"categoria": "geral", "titulo": "Hack"})
        self.assertEqual(Dica.objects.count(), antes)
        self.client.post(reverse("dica_delete", args=[d.id]))
        self.assertTrue(Dica.objects.filter(id=d.id).exists())

    def test_dashboard_renders_seeded(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("dicas_dashboard"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "dica-card")


class StarlinkTests(TestCase):
    """Modulo Starlinks: CRUD, senha (manter na edicao) e permissoes.

    Obs.: o banco de teste ja vem com as Starlinks do seed (migration 0027).
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def test_ti_creates_starlink(self):
        self.client.force_login(self.ti)
        antes = Starlink.objects.count()
        resp = self.client.post(
            reverse("starlink_create"),
            {
                "nome": "Star99", "local": "Obra X", "email": "star99@sidertec.com.br",
                "ativo": "1", "forma_pagamento": "cartao",
                "final_cartao": "1234", "numero_serie": "SN123", "numero_kit": "KIT123",
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Starlink.objects.count(), antes + 1)
        s = Starlink.objects.get(nome="Star99")
        self.assertTrue(s.ativo)
        self.assertEqual(s.numero_serie, "SN123")
        self.assertEqual(s.criado_por, self.ti)

    def test_create_requires_name(self):
        self.client.force_login(self.ti)
        antes = Starlink.objects.count()
        self.client.post(reverse("starlink_create"), {"nome": "", "email": "x@x.com"})
        self.assertEqual(Starlink.objects.count(), antes)

    def test_inactive_when_unchecked(self):
        self.client.force_login(self.ti)
        self.client.post(reverse("starlink_create"), {"nome": "Inativa", "forma_pagamento": "pix"})
        s = Starlink.objects.get(nome="Inativa")
        self.assertFalse(s.ativo)  # sem 'ativo' no POST => inativa
        self.assertEqual(s.forma_pagamento, "pix")

    def test_update_changes_fields(self):
        self.client.force_login(self.ti)
        s = Starlink.objects.create(nome="Star", local="Antigo", ativo=True)
        self.client.post(
            reverse("starlink_update", args=[s.id]),
            {"nome": "Star", "local": "Novo", "ativo": "1", "forma_pagamento": "cartao"},
        )
        s.refresh_from_db()
        self.assertEqual(s.local, "Novo")

    def test_delete(self):
        self.client.force_login(self.ti)
        s = Starlink.objects.create(nome="Apagar")
        resp = self.client.post(reverse("starlink_delete", args=[s.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Starlink.objects.filter(id=s.id).exists())

    def test_common_user_blocked(self):
        s = Starlink.objects.create(nome="Protegida")
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("starlinks_dashboard")).status_code, 302)
        antes = Starlink.objects.count()
        self.client.post(reverse("starlink_create"), {"nome": "Hack"})
        self.assertEqual(Starlink.objects.count(), antes)
        self.client.post(reverse("starlink_delete", args=[s.id]))
        self.assertTrue(Starlink.objects.filter(id=s.id).exists())

    def test_dashboard_renders_seeded(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("starlinks_dashboard"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "star-card")


class CofreCryptoTests(TestCase):
    """Cifra/decifra das credenciais do cofre."""

    def test_encrypt_decrypt_round_trip(self):
        from core.crypto import decrypt_text, encrypt_text
        token = encrypt_text("segredo-super-secreto")
        self.assertNotEqual(token, "segredo-super-secreto")  # cifrado
        self.assertEqual(decrypt_text(token), "segredo-super-secreto")
        self.assertEqual(decrypt_text(""), "")

    def test_credencial_set_get_password(self):
        c = CofreCredencial(rotulo="X")
        c.definir_senha("minha-senha")
        self.assertNotIn("minha-senha", c.senha_cifrada)  # nao fica em texto
        self.assertEqual(c.obter_senha(), "minha-senha")


class CofreTests(TestCase):
    """Cofre: senha-mestra, destrave, revelar sob demanda, auditoria e permissoes."""

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        self.admin = User.objects.create_user(username="adm", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        Group.objects.get_or_create(name=ADMIN_GROUP_NAME)
        att = Group.objects.get(name=ATTENDANT_GROUP_NAME)
        adm = Group.objects.get(name=ADMIN_GROUP_NAME)
        self.ti.groups.add(att)
        self.admin.groups.add(adm, att)
        # Zera qualquer config de seed para comecar do estado 'setup'.
        CofreConfig.objects.all().delete()

    def test_dashboard_setup_state_without_master(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("cofre_dashboard"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Configurar o cofre")

    def test_common_user_blocked(self):
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("cofre_dashboard")).status_code, 302)

    def test_only_admin_sets_master(self):
        # TI (nao admin) nao pode definir senha-mestra
        self.client.force_login(self.ti)
        self.client.post(reverse("cofre_set_master"), {"nova_senha": "abcdef", "confirma_senha": "abcdef"})
        self.assertFalse(CofreConfig.load().tem_senha_mestra)
        # Admin pode
        self.client.force_login(self.admin)
        self.client.post(reverse("cofre_set_master"), {"nova_senha": "abcdef", "confirma_senha": "abcdef"})
        self.assertTrue(CofreConfig.load().tem_senha_mestra)

    def test_unlock_flow_and_reveal_requires_unlock(self):
        cfg = CofreConfig.load()
        cfg.definir_senha_mestra("chave123")
        cfg.save()
        cred = CofreCredencial(rotulo="Roteador", usuario="admin")
        cred.definir_senha("root123")
        cred.save()

        self.client.force_login(self.ti)
        # Travado: revelar deve dar 403
        r = self.client.post(reverse("cofre_credencial_reveal", args=[cred.id]))
        self.assertEqual(r.status_code, 403)

        # Senha-mestra errada
        self.client.post(reverse("cofre_unlock"), {"senha_mestra": "errada"})
        r = self.client.post(reverse("cofre_credencial_reveal", args=[cred.id]))
        self.assertEqual(r.status_code, 403)

        # Senha-mestra certa -> destrava
        self.client.post(reverse("cofre_unlock"), {"senha_mestra": "chave123"})
        r = self.client.post(reverse("cofre_credencial_reveal", args=[cred.id]))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json()["senha"], "root123")
        # Revelacao auditada
        self.assertTrue(CofreAuditoria.objects.filter(acao=CofreAuditoria.ACAO_CRED_REVELADA, credencial=cred).exists())

    def test_lockout_after_max_attempts(self):
        cfg = CofreConfig.load()
        cfg.definir_senha_mestra("chave123")
        cfg.save()
        self.client.force_login(self.ti)
        for _ in range(5):
            self.client.post(reverse("cofre_unlock"), {"senha_mestra": "errada"})
        cfg.refresh_from_db()
        self.assertTrue(cfg.esta_bloqueado())
        # Mesmo com a senha certa, bloqueado nao destrava
        self.client.post(reverse("cofre_unlock"), {"senha_mestra": "chave123"})
        r = self.client.post(reverse("cofre_credencial_reveal", args=[CofreCredencial.objects.create(rotulo="Y").id]))
        self.assertEqual(r.status_code, 403)

    def test_credential_crud_requires_unlock(self):
        cfg = CofreConfig.load()
        cfg.definir_senha_mestra("chave123")
        cfg.save()
        self.client.force_login(self.ti)
        # Travado: criar credencial nao funciona
        antes = CofreCredencial.objects.count()
        self.client.post(reverse("cofre_credencial_create"), {"rotulo": "Nova", "senha": "s"})
        self.assertEqual(CofreCredencial.objects.count(), antes)
        # Destrava e cria
        self.client.post(reverse("cofre_unlock"), {"senha_mestra": "chave123"})
        self.client.post(reverse("cofre_credencial_create"), {"rotulo": "Nova", "usuario": "u", "senha": "s3nha"})
        cred = CofreCredencial.objects.get(rotulo="Nova")
        self.assertEqual(cred.obter_senha(), "s3nha")
        # Atualiza mantendo senha (branco)
        self.client.post(reverse("cofre_credencial_update", args=[cred.id]), {"rotulo": "Nova2", "senha": ""})
        cred.refresh_from_db()
        self.assertEqual(cred.rotulo, "Nova2")
        self.assertEqual(cred.obter_senha(), "s3nha")
        # Exclui
        self.client.post(reverse("cofre_credencial_delete", args=[cred.id]))
        self.assertFalse(CofreCredencial.objects.filter(id=cred.id).exists())

    def test_lock_clears_session(self):
        cfg = CofreConfig.load()
        cfg.definir_senha_mestra("chave123")
        cfg.save()
        self.client.force_login(self.ti)
        self.client.post(reverse("cofre_unlock"), {"senha_mestra": "chave123"})
        cred = CofreCredencial.objects.create(rotulo="Z")
        self.assertEqual(self.client.post(reverse("cofre_credencial_reveal", args=[cred.id])).status_code, 200)
        self.client.post(reverse("cofre_lock"))
        self.assertEqual(self.client.post(reverse("cofre_credencial_reveal", args=[cred.id])).status_code, 403)


_LOCMEM = "django.core.mail.backends.locmem.EmailBackend"


@override_settings(MEDIA_ROOT=tempfile.mkdtemp(), EMAIL_BACKEND_OVERRIDE=_LOCMEM)
class EmailNotificacaoTests(TestCase):
    """Notificacoes por e-mail: disparo nos eventos e configuracao (SMTP)."""

    def setUp(self):
        from .models import EmailConfig

        User = get_user_model()
        self.owner = User.objects.create_user(username="joao", password="x", email="joao@empresa.com")
        self.attendant = User.objects.create_user(username="ti", password="x", email="atendente@empresa.com")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.attendant.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

        self.config = EmailConfig.load()
        self.config.ativo = True
        self.config.usuario = "chamados@empresa.com"
        self.config.remetente = "chamados@empresa.com"
        self.config.emails_ti = "suporte@empresa.com, ti@empresa.com"
        self.config.save()

    # --- Disparo dos eventos ---------------------------------------------
    def test_novo_chamado_portal_notifica_solicitante_e_ti(self):
        from django.core import mail

        mail.outbox = []
        self.client.force_login(self.owner)
        resp = self.client.post(
            reverse("open_ticket"), {"titulo": "PC nao liga", "descricao": "O computador nao liga de jeito nenhum"}
        )
        self.assertEqual(resp.status_code, 302)
        destinos = {d for m in mail.outbox for d in m.to}
        self.assertIn("joao@empresa.com", destinos)
        self.assertIn("suporte@empresa.com", destinos)
        self.assertIn("ti@empresa.com", destinos)

    def test_solicitante_da_ti_nao_recebe_copia_pessoal(self):
        # Quando o solicitante e da propria TI (ex.: atendente abre/converte um
        # chamado para si), ele so recebe pela lista da TI, sem a confirmacao
        # pessoal — evitando dois e-mails da mesma acao.
        from django.core import mail
        from .emails import notificar_novo_chamado

        chamado = Chamado.objects.create(
            numero="CH-000210", titulo="Meu proprio chamado", solicitante=self.attendant,
            solicitante_nome="TI", solicitante_email="atendente@empresa.com",
            status=Chamado.STATUS_ABERTO,
        )
        mail.outbox = []
        notificar_novo_chamado(chamado)
        destinos = {d for m in mail.outbox for d in m.to}
        # Recebe apenas pela lista da TI; nada de copia pessoal (um unico e-mail).
        self.assertNotIn("atendente@empresa.com", destinos)
        self.assertIn("suporte@empresa.com", destinos)
        self.assertIn("ti@empresa.com", destinos)

    def test_notificacoes_desligadas_nao_enviam(self):
        from django.core import mail

        self.config.ativo = False
        self.config.save()
        mail.outbox = []
        self.client.force_login(self.owner)
        resp = self.client.post(
            reverse("open_ticket"),
            {"titulo": "Sem rede", "descricao": "Nao tenho acesso a internet aqui"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(len(mail.outbox), 0)

    def test_mensagem_do_solicitante_notifica_so_ti(self):
        from django.core import mail

        chamado = Chamado.objects.create(
            numero="CH-000201", titulo="T", solicitante=self.owner,
            solicitante_email="joao@empresa.com", status=Chamado.STATUS_ABERTO,
        )
        mail.outbox = []
        self.client.force_login(self.owner)
        resp = self.client.post(
            reverse("ticket_message_create", args=[chamado.numero]), {"texto": "Ola"}
        )
        self.assertEqual(resp.status_code, 302)
        destinos = {d for m in mail.outbox for d in m.to}
        # O proprio autor (solicitante) nao recebe copia.
        self.assertNotIn("joao@empresa.com", destinos)
        self.assertIn("suporte@empresa.com", destinos)

    def test_mensagem_da_ti_notifica_solicitante(self):
        from django.core import mail

        chamado = Chamado.objects.create(
            numero="CH-000202", titulo="T", solicitante=self.owner,
            solicitante_email="joao@empresa.com", status=Chamado.STATUS_EM_ATENDIMENTO,
        )
        mail.outbox = []
        self.client.force_login(self.attendant)
        self.client.post(reverse("ticket_message_create", args=[chamado.numero]), {"texto": "resolvendo"})
        destinos = {d for m in mail.outbox for d in m.to}
        self.assertIn("joao@empresa.com", destinos)
        self.assertNotIn("atendente@empresa.com", destinos)

    def test_mudanca_status_notifica(self):
        from django.core import mail

        chamado = Chamado.objects.create(
            numero="CH-000203", titulo="T", solicitante=self.owner,
            solicitante_email="joao@empresa.com", status=Chamado.STATUS_ABERTO,
        )
        mail.outbox = []
        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("move_ticket"),
            data=json.dumps(
                {"ticket_number": chamado.numero, "target": "atendente", "attendant_id": self.attendant.id}
            ),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertGreaterEqual(len(mail.outbox), 1)
        assunto = " ".join(m.subject for m in mail.outbox)
        self.assertIn(chamado.numero, assunto)

    def test_fechamento_stop_notifica(self):
        from django.core import mail

        chamado = Chamado.objects.create(
            numero="CH-000204", titulo="T", solicitante=self.owner,
            solicitante_email="joao@empresa.com", status=Chamado.STATUS_EM_ATENDIMENTO,
            atendente_atual=self.attendant,
        )
        AtendimentoHistorico.objects.create(
            chamado=chamado, atendente=self.attendant, iniciado_em=timezone.now()
        )
        mail.outbox = []
        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("finish_attendance"),
            data=json.dumps({"ticket_number": chamado.numero, "action": "stop", "description": "trocado o cabo"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        corpos = " ".join(m.body for m in mail.outbox)
        self.assertIn("trocado o cabo", corpos)
        self.assertIn("joao@empresa.com", {d for m in mail.outbox for d in m.to})

    def test_falha_de_envio_nao_quebra_o_chamado(self):
        # Sem remetente configurado, o envio e ignorado silenciosamente.
        self.config.usuario = ""
        self.config.remetente = ""
        self.config.save()
        self.client.force_login(self.owner)
        resp = self.client.post(
            reverse("open_ticket"),
            {"titulo": "Teste sem remetente", "descricao": "Deve abrir mesmo sem e-mail"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(Chamado.objects.filter(titulo="Teste sem remetente").exists())

    # --- Tela de configuracao --------------------------------------------
    def test_config_screen_ti_ok_comum_bloqueado(self):
        self.client.force_login(self.attendant)
        self.assertEqual(self.client.get(reverse("email_config")).status_code, 200)

        comum = get_user_model().objects.create_user(username="c", password="x")
        self.client.force_login(comum)
        resp = self.client.get(reverse("email_config"))
        self.assertEqual(resp.status_code, 302)

    def test_salvar_config_cifra_a_senha(self):
        from .models import EmailConfig

        self.client.force_login(self.attendant)
        resp = self.client.post(
            reverse("email_config_save"),
            {
                "ativo": "on", "host": "smtp.gmail.com", "porta": "587", "usar_tls": "on",
                "timeout": "15", "usuario": "conta@gmail.com", "remetente_nome": "Chamados TI",
                "emails_ti": "ti@empresa.com", "senha": "abcd efgh ijkl mnop",
                "notif_novo_chamado": "on",
            },
        )
        self.assertEqual(resp.status_code, 302)
        cfg = EmailConfig.load()
        self.assertTrue(cfg.tem_senha)
        # Espacos removidos e senha recuperavel (cifrada em repouso).
        self.assertEqual(cfg.obter_senha(), "abcdefghijklmnop")
        # A senha em texto nao aparece no campo cifrado do banco.
        self.assertNotIn("abcdefghijklmnop", cfg.senha_cifrada)

    def test_salvar_sem_senha_mantem_a_atual(self):
        from .models import EmailConfig

        self.config.definir_senha("segredo123")
        self.config.save()
        self.client.force_login(self.attendant)
        self.client.post(
            reverse("email_config_save"),
            {"ativo": "on", "host": "smtp.gmail.com", "porta": "587", "usar_tls": "on",
             "usuario": "conta@gmail.com", "emails_ti": "ti@empresa.com", "senha": ""},
        )
        self.assertEqual(EmailConfig.load().obter_senha(), "segredo123")

    def test_comum_nao_salva_config(self):
        comum = get_user_model().objects.create_user(username="c2", password="x")
        self.client.force_login(comum)
        resp = self.client.post(reverse("email_config_save"), {"host": "x"})
        self.assertEqual(resp.status_code, 302)  # redirecionado, sem salvar

    def test_email_de_teste(self):
        from django.core import mail

        mail.outbox = []
        self.client.force_login(self.attendant)
        resp = self.client.post(reverse("email_config_test"), {"email_teste": "quem@empresa.com"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("quem@empresa.com", mail.outbox[0].to)

    def test_tls_e_ssl_juntos_sao_rejeitados(self):
        from .models import EmailConfig

        self.client.force_login(self.attendant)
        self.client.post(
            reverse("email_config_save"),
            {"host": "smtp.gmail.com", "porta": "587", "usar_tls": "on", "usar_ssl": "on",
             "usuario": "c@g.com", "emails_ti": "ti@e.com"},
        )
        # Nao deve ter ligado ssl junto com tls (rejeitado).
        cfg = EmailConfig.load()
        self.assertFalse(cfg.usar_tls and cfg.usar_ssl)


class InsumoUpdateTests(TestCase):
    """Edicao de insumo (ajuste de estoque) do modulo Insumos."""

    def setUp(self):
        from .models import InsumoTI

        User = get_user_model()
        self.ti = User.objects.create_user(username="ti_ins", password="x")
        self.common = User.objects.create_user(username="comum_ins", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))
        self.insumo = InsumoTI.objects.create(nome="Bateria", quantidade_atual=0)

    def _post(self, name, data):
        return self.client.post(
            reverse(name, args=[self.insumo.id]),
            data=json.dumps(data),
            content_type="application/json",
        )

    def test_editar_nao_altera_quantidade(self):
        self.insumo.quantidade_atual = 10
        self.insumo.save()
        self.client.force_login(self.ti)
        resp = self._post("insumo_update", {"nome": "Bateria AA", "descricao": "recarregavel", "observacao": ""})
        self.assertEqual(resp.status_code, 200)
        self.insumo.refresh_from_db()
        self.assertEqual(self.insumo.nome, "Bateria AA")
        self.assertEqual(self.insumo.descricao, "recarregavel")
        self.assertEqual(self.insumo.quantidade_atual, 10)  # inalterada pela edicao

    def test_entrada_soma_ao_estoque(self):
        self.insumo.quantidade_atual = 5
        self.insumo.save()
        self.client.force_login(self.ti)
        resp = self._post("insumo_entrada", {"quantidade": 8})
        self.assertEqual(resp.status_code, 200)
        self.insumo.refresh_from_db()
        self.assertEqual(self.insumo.quantidade_atual, 13)

    def test_entrada_invalida_rejeitada(self):
        self.client.force_login(self.ti)
        self.assertEqual(self._post("insumo_entrada", {"quantidade": 0}).status_code, 400)
        self.assertEqual(self._post("insumo_entrada", {"quantidade": -3}).status_code, 400)

    def test_entrada_registra_movimento_no_extrato(self):
        from .models import RetiradaInsumoTI

        self.client.force_login(self.ti)
        self._post("insumo_entrada", {"quantidade": 3})
        self.assertTrue(
            RetiradaInsumoTI.objects.filter(
                insumo=self.insumo, tipo=RetiradaInsumoTI.TIPO_ENTRADA, quantidade=3
            ).exists()
        )

    def test_excluir_insumo(self):
        from .models import InsumoTI

        self.client.force_login(self.ti)
        resp = self._post("insumo_delete", {})
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(InsumoTI.objects.filter(id=self.insumo.id).exists())

    def test_comum_nao_edita_nem_entrada_nem_exclui(self):
        self.client.force_login(self.common)
        self.assertEqual(self._post("insumo_update", {"nome": "X"}).status_code, 403)
        self.assertEqual(self._post("insumo_entrada", {"quantidade": 5}).status_code, 403)
        self.assertEqual(self._post("insumo_delete", {}).status_code, 403)

    def test_busca_retiradas(self):
        from .models import RetiradaInsumoTI

        RetiradaInsumoTI.objects.create(
            insumo=self.insumo, quantidade=1, entregue_para="Joao Silva", motivo="Departamento: RH")
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("retiradas_search"), {"q": "joao"})
        self.assertEqual(resp.status_code, 200)
        res = resp.json()["resultados"]
        self.assertTrue(any("Joao" in x["entregue_para"] for x in res))
        # termo sem correspondencia
        vazio = self.client.get(reverse("retiradas_search"), {"q": "zzznaoexiste"})
        self.assertEqual(len(vazio.json()["resultados"]), 0)

    def test_busca_retiradas_comum_bloqueado(self):
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("retiradas_search")).status_code, 403)


class NotificacoesStreamTests(TestCase):
    """Endpoint SSE de notificacoes (permissao e cabecalhos). Nao consome o
    stream (loop infinito) — valida apenas o handshake."""

    def setUp(self):
        User = get_user_model()
        self.ti = User.objects.create_user(username="ti_sse", password="x")
        self.common = User.objects.create_user(username="comum_sse", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def test_ti_recebe_event_stream(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("notificacoes_stream"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/event-stream", resp["Content-Type"])
        self.assertEqual(resp["X-Accel-Buffering"], "no")

    def test_comum_bloqueado(self):
        self.client.force_login(self.common)
        resp = self.client.get(reverse("notificacoes_stream"))
        self.assertEqual(resp.status_code, 403)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class EmprestimoEdicaoTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))

    def _novo_emprestimo(self):
        emp = EmprestimoTI.objects.create(
            colaborador_nome="Fulano",
            data_emprestimo="2026-01-10",
            status=EmprestimoTI.STATUS_ASSINADA_OK,
            termo_assinado_ok=True,
            criado_por=self.ti,
        )
        equip = EquipamentoEmprestimoTI.objects.create(
            emprestimo=emp, tipo_equipamento="Notebook", data_emprestimo="2026-01-10"
        )
        return emp, equip

    def test_adicionar_equipamento_volta_para_aguardando(self):
        emp, equip = self._novo_emprestimo()
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("emprestimo_editar", args=[emp.id]),
            {
                "colaborador_nome": "Fulano",
                "data_emprestimo": "2026-01-10",
                f"acao_equip_{equip.id}": "manter",
                "equipamentos_count": "1",
                "equip_0_tipo": "Monitor",
                "equip_0_data": "2026-07-05",
            },
        )
        self.assertEqual(resp.status_code, 200)
        emp.refresh_from_db()
        self.assertEqual(emp.equipamentos.count(), 2)
        self.assertEqual(emp.status, EmprestimoTI.STATUS_AGUARDANDO)
        self.assertFalse(emp.termo_assinado_ok)
        self.assertTrue(emp.termo_pdf)
        novo = emp.equipamentos.get(tipo_equipamento="Monitor")
        self.assertEqual(novo.data_emprestimo.isoformat(), "2026-07-05")

    def test_devolver_unico_equipamento_marca_devolvido(self):
        emp, equip = self._novo_emprestimo()
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("emprestimo_editar", args=[emp.id]),
            {
                "colaborador_nome": "Fulano",
                "data_emprestimo": "2026-01-10",
                f"acao_equip_{equip.id}": "devolver",
                f"devolver_data_{equip.id}": "2026-07-08",
                "equipamentos_count": "0",
            },
        )
        self.assertEqual(resp.status_code, 200)
        equip.refresh_from_db()
        emp.refresh_from_db()
        self.assertEqual(equip.data_devolucao.isoformat(), "2026-07-08")
        self.assertEqual(emp.status, EmprestimoTI.STATUS_DEVOLVIDO)

    def test_nao_remove_todos_os_equipamentos(self):
        emp, equip = self._novo_emprestimo()
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("emprestimo_editar", args=[emp.id]),
            {
                "colaborador_nome": "Fulano",
                "data_emprestimo": "2026-01-10",
                f"acao_equip_{equip.id}": "remover",
                "equipamentos_count": "0",
            },
        )
        self.assertEqual(resp.status_code, 400)
        emp.refresh_from_db()
        self.assertEqual(emp.equipamentos.count(), 1)

    def test_comum_bloqueado(self):
        emp, equip = self._novo_emprestimo()
        self.client.force_login(self.common)
        resp = self.client.post(
            reverse("emprestimo_editar", args=[emp.id]),
            {"colaborador_nome": "Fulano", "data_emprestimo": "2026-01-10", "equipamentos_count": "0"},
        )
        self.assertEqual(resp.status_code, 403)


class PlanilhaAtendimentosTests(TestCase):
    """Planilha mensal de atendimentos por atendente (modelo .xlsx da TI).

    Regra central: uma linha por periodo Play -> Pause/Stop que COMECOU no mes.
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(
            username="ti", password="x", first_name="Fabiano", last_name="Polone", email="fabiano@x.com"
        )
        self.outro = User.objects.create_user(username="outro.ti", password="x", first_name="Marina")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        grupo = Group.objects.get(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(grupo)
        self.outro.groups.add(grupo)

        self.solicitante = User.objects.create_user(
            username="dandara", password="x", first_name="Dandara", last_name="Santiago", email="dandara@x.com"
        )
        Ramal.objects.create(colaborador="Dandara Santiago", setor="RH", email="dandara@x.com", telefone="1234")
        Ramal.objects.create(
            colaborador="Fabiano Polone", setor="TI", email="fabiano@x.com", telefone="(14) 98820-8134"
        )

    def _chamado(self, titulo, origem="Portal do solicitante", solicitante=None):
        return Chamado.objects.create(
            numero=Chamado.gerar_numero(),
            titulo=titulo,
            solicitante=solicitante or self.solicitante,
            atendente_atual=self.ti,
            origem=origem,
            status=Chamado.STATUS_ATRIBUIDO,
        )

    def _periodo(self, chamado, inicio, fim, descricao="Feito", atendente=None):
        return AtendimentoHistorico.objects.create(
            chamado=chamado,
            atendente=atendente or self.ti,
            iniciado_em=inicio,
            finalizado_em=fim,
            duracao=(fim - inicio) if fim else None,
            tipo_encerramento=AtendimentoHistorico.TIPO_ENCERRAMENTO_STOP if fim else "",
            descricao_atividade=descricao,
        )

    def _abrir(self, resposta):
        import io as _io

        import openpyxl

        return openpyxl.load_workbook(_io.BytesIO(resposta.content))

    def _dt(self, dia, hora, minuto=0, mes=5, ano=2026):
        from datetime import datetime

        return timezone.make_aware(datetime(ano, mes, dia, hora, minuto))

    def _baixar(self, atendente=None, mes="2026-05"):
        return self.client.get(reverse("atendimentos_planilha", args=[(atendente or self.ti).id]), {"mes": mes})

    # ----- uma linha por periodo -----
    def test_cada_play_stop_gera_uma_linha(self):
        chamado = self._chamado("Fazer melhoria sistema chamados")
        self._periodo(chamado, self._dt(5, 8, 12), self._dt(5, 17, 45), "Feito uma parte")
        self._periodo(chamado, self._dt(6, 8, 20), self._dt(6, 10, 0), "Terminado")

        self.client.force_login(self.ti)
        resp = self._baixar()
        self.assertEqual(resp.status_code, 200)
        ws = self._abrir(resp).active

        # Duas linhas (8 e 9) para o mesmo chamado, uma por periodo trabalhado.
        self.assertEqual(ws["E8"].value, "Fazer melhoria sistema chamados")
        self.assertEqual(ws["E9"].value, "Fazer melhoria sistema chamados")
        self.assertIsNone(ws["E10"].value)
        self.assertEqual(ws["H8"].value, "Feito uma parte")
        self.assertEqual(ws["H9"].value, "Terminado")
        # Data = Play, Fechado = Stop, Tempo = formula
        self.assertEqual(ws["B8"].value.strftime("%d/%m/%Y %H:%M"), "05/05/2026 08:12")
        self.assertEqual(ws["I8"].value.strftime("%d/%m/%Y %H:%M"), "05/05/2026 17:45")
        # O Tempo vai calculado (nao como formula): o openpyxl grava formula
        # sem valor em cache e o Excel abria a celula em branco.
        self.assertEqual(ws["J8"].value, timezone.timedelta(hours=9, minutes=33))
        self.assertEqual(ws["J8"].number_format, "[h]:mm:ss")

    def test_colunas_contato_setor_falha_e_tk(self):
        chamado = self._chamado("Preciso de um fone")
        self._periodo(chamado, self._dt(5, 10, 20), self._dt(5, 10, 41), "Disponibilizado")

        self.client.force_login(self.ti)
        ws = self._abrir(self._baixar()).active
        self.assertIsNone(ws["A8"].value)  # Tk fica vazio, como nas planilhas atuais
        self.assertEqual(ws["C8"].value, "Dandara Santiago")
        self.assertEqual(ws["D8"].value, "RH")  # setor vindo do Ramal (casado pelo e-mail)
        self.assertEqual(ws["F8"].value, "Baixa")
        self.assertEqual(ws["G8"].value, "N/A")
        self.assertIsNone(ws["K8"].value)  # Acao Eficaz e preenchida a mao

    def test_chamado_criado_pela_ti_entra_como_programada(self):
        chamado = self._chamado("Fazer orcamentos", origem="Kanban TI", solicitante=self.ti)
        self._periodo(chamado, self._dt(5, 15, 44), self._dt(5, 17, 45))
        pendencia = self._chamado("Cadastrar contratos", origem="Pendencia TI", solicitante=self.ti)
        self._periodo(pendencia, self._dt(6, 9, 0), self._dt(6, 9, 30))

        self.client.force_login(self.ti)
        ws = self._abrir(self._baixar()).active
        self.assertEqual(ws["F8"].value, "Programada")
        self.assertEqual(ws["F9"].value, "Programada")
        self.assertEqual(ws["D8"].value, "TI")

    def test_cabecalho_traz_mes_atendente_e_telefone(self):
        self.client.force_login(self.ti)
        ws = self._abrir(self._baixar()).active
        self.assertEqual(ws["A4"].value, "Atendimentos TI Sidertec - 05/2026")
        self.assertEqual(ws["A5"].value, "Fabiano Polone (14) 98820-8134")
        self.assertEqual(ws.title, "Maio")
        self.assertEqual(ws["B7"].value, "Data")  # cabecalho do modelo preservado

    def test_so_entram_periodos_do_mes_e_do_atendente(self):
        chamado = self._chamado("Chamado do mes")
        self._periodo(chamado, self._dt(5, 9, 0), self._dt(5, 10, 0), "Do mes")
        self._periodo(chamado, self._dt(20, 9, 0, mes=4), self._dt(20, 10, 0, mes=4), "Abril")
        self._periodo(chamado, self._dt(7, 9, 0), self._dt(7, 10, 0), "Da Marina", atendente=self.outro)

        self.client.force_login(self.ti)
        ws = self._abrir(self._baixar()).active
        self.assertEqual(ws["H8"].value, "Do mes")
        self.assertIsNone(ws["H9"].value)

    def test_periodo_em_andamento_sai_sem_fechado_e_sem_tempo(self):
        chamado = self._chamado("Play aberto")
        self._periodo(chamado, self._dt(5, 9, 0), None, "")

        self.client.force_login(self.ti)
        ws = self._abrir(self._baixar()).active
        self.assertIsNotNone(ws["B8"].value)
        self.assertIsNone(ws["I8"].value)
        self.assertIsNone(ws["J8"].value)

    def test_encerramento_direto_sem_play_nao_gera_linha(self):
        # Stop de chamado em "aguardando" nao cria AtendimentoHistorico: por
        # decisao de uso, nao entra na planilha.
        chamado = self._chamado("Aguardando peca")
        chamado.status = Chamado.STATUS_AGUARDANDO_PECA
        chamado.save(update_fields=["status"])
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("finish_attendance"),
            data=json.dumps({"ticket_number": chamado.numero, "action": "stop", "description": "Peca nao veio"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(AtendimentoHistorico.objects.filter(chamado=chamado).count(), 0)

        hoje = timezone.localdate()
        ws = self._abrir(self._baixar(mes=f"{hoje.year}-{hoje.month:02d}")).active
        self.assertIsNone(ws["E8"].value)

    def test_nome_do_arquivo_segue_o_padrao_da_ti(self):
        self.client.force_login(self.ti)
        resp = self._baixar()
        self.assertIn('filename="05-2026 - Fabiano.xlsx"', resp["Content-Disposition"])
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_qualquer_atendente_baixa_de_qualquer_atendente(self):
        self.client.force_login(self.outro)
        self.assertEqual(self._baixar(self.ti).status_code, 200)

    def test_usuario_comum_nao_baixa(self):
        self.client.force_login(self.common)
        self.assertEqual(self._baixar().status_code, 302)

    def test_mes_invalido_retorna_400_e_usuario_nao_atendente_404(self):
        self.client.force_login(self.ti)
        self.assertEqual(self._baixar(mes="abc").status_code, 400)
        self.assertEqual(self._baixar(mes="2026-13").status_code, 400)
        self.assertEqual(
            self.client.get(reverse("atendimentos_planilha", args=[self.solicitante.id])).status_code, 404
        )

    def test_botao_e_modal_aparecem_no_kanban(self):
        self.client.force_login(self.ti)
        html = self.client.get(reverse("tickets_dashboard")).content.decode()
        self.assertIn(f'data-planilha-atendente="{self.ti.id}"', html)
        self.assertIn('id="planilhaAtendimentosModal"', html)
        self.assertIn('id="planilhaMes"', html)

    def test_resumo_do_cabecalho_vem_calculado(self):
        # O modelo traz COUNTIF/SUM, mas o openpyxl grava formula sem valor em
        # cache: o Excel abria o bloco em branco e o grafico lia zero.
        for i in range(3):
            self._periodo(self._chamado(f"Do usuario {i}"), self._dt(5, 9 + i, 0), self._dt(5, 10 + i, 0))
        interno = self._chamado("Tarefa da TI", origem="Kanban TI", solicitante=self.ti)
        self._periodo(interno, self._dt(6, 9, 0), self._dt(6, 10, 0))

        self.client.force_login(self.ti)
        ws = self._abrir(self._baixar()).active
        self.assertEqual(ws["E2"].value, 4)  # total de linhas
        self.assertEqual(ws["F2"].value, 0)  # Alta
        self.assertEqual(ws["F3"].value, 0)  # Media
        self.assertEqual(ws["F4"].value, 3)  # Baixa (usuario)
        self.assertEqual(ws["F5"].value, 1)  # Programada (TI)
        # os rotulos do modelo continuam na coluna G
        self.assertEqual(ws["G4"].value, "Baixa")
        self.assertEqual(ws["G5"].value, "Programada")

    def test_atendimento_de_mais_de_24h_nao_da_a_volta_no_relogio(self):
        # Existem periodos reais de dias (Play esquecido aberto): com o formato
        # do modelo ("h:mm:ss AM/PM") 169h apareceriam como "1:37 AM".
        chamado = self._chamado("Play esquecido aberto")
        self._periodo(chamado, self._dt(5, 8, 0), self._dt(12, 9, 30))

        self.client.force_login(self.ti)
        ws = self._abrir(self._baixar()).active
        self.assertEqual(ws["J8"].value, timezone.timedelta(days=7, hours=1, minutes=30))
        self.assertEqual(ws["J8"].number_format, "[h]:mm:ss")

    def test_modal_oferece_apenas_meses_com_atendimento(self):
        # O sistema so tem historico desde que o controle de tempo entrou em uso:
        # oferecer meses fixos faria o usuario baixar planilhas em branco.
        chamado = self._chamado("Chamado com atendimento")
        self._periodo(chamado, self._dt(5, 9, 0), self._dt(5, 10, 0))
        self._periodo(chamado, self._dt(20, 9, 0, mes=4), self._dt(20, 10, 0, mes=4))

        self.client.force_login(self.ti)
        colunas = self.client.get(reverse("tickets_dashboard")).context["attendant_columns"]
        minha = next(c for c in colunas if c["attendant_id"] == self.ti.id)
        meses = json.loads(minha["planilha_meses_json"])
        valores = [m["valor"] for m in meses]

        hoje = timezone.localdate()
        atual = f"{hoje.year}-{hoje.month:02d}"
        # Os dois meses com atendimento + o mes atual (sempre disponivel).
        self.assertIn("2026-05", valores)
        self.assertIn("2026-04", valores)
        self.assertIn(atual, valores)
        self.assertNotIn("2026-03", valores)  # mes sem atendimento nao e oferecido
        self.assertEqual(valores[0], atual)  # mais recente primeiro (padrao)
        por_valor = {m["valor"]: m["total"] for m in meses}
        self.assertEqual(por_valor["2026-05"], 1)
        self.assertEqual(por_valor["2026-04"], 1)

    def test_atendente_sem_atendimento_recebe_apenas_o_mes_atual(self):
        self.client.force_login(self.ti)
        colunas = self.client.get(reverse("tickets_dashboard")).context["attendant_columns"]
        dele = next(c for c in colunas if c["attendant_id"] == self.outro.id)
        meses = json.loads(dele["planilha_meses_json"])
        hoje = timezone.localdate()
        self.assertEqual([m["valor"] for m in meses], [f"{hoje.year}-{hoje.month:02d}"])
        self.assertEqual(meses[0]["total"], 0)


class ImportaAtendimentosLegadoTests(TestCase):
    """Importacao dos periodos de atendimento do sistema antigo (ERP-TI).

    O foco dos testes sao as protecoes: nao encostar em chamado recente, nunca
    criar periodo sem fim (que viraria atendimento ativo no Kanban) e nao
    duplicar o que o sistema novo ja registrou.
    """

    ORIGEM = "Migrado (sistema antigo)"

    def setUp(self):
        User = get_user_model()
        self.ti = User.objects.create_user(username="fabiano.polone", password="x", first_name="Fabiano")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))
        self.solicitante = User.objects.create_user(username="usuario", password="x")

    # ----- banco legado sintetico -----
    def _banco_legado(self, tickets, periodos, usuarios=None):
        """Cria um sqlite com o schema minimo do sistema antigo."""
        import sqlite3
        import tempfile

        caminho = os.path.join(tempfile.mkdtemp(), "legado.sqlite3")
        con = sqlite3.connect(caminho)
        con.execute("CREATE TABLE auth_user (id INTEGER PRIMARY KEY, username TEXT)")
        con.execute("CREATE TABLE chamados_ticket (id INTEGER PRIMARY KEY, created_at TEXT)")
        con.execute(
            "CREATE TABLE chamados_ticketattendance ("
            " id INTEGER PRIMARY KEY, ticket_id INTEGER, attendant_id INTEGER,"
            " started_at TEXT, ended_at TEXT, end_action TEXT, note TEXT)"
        )
        for uid, username in (usuarios or [(1, "fabiano.polone")]):
            con.execute("INSERT INTO auth_user VALUES (?, ?)", (uid, username))
        for tid, criado in tickets:
            con.execute("INSERT INTO chamados_ticket VALUES (?, ?)", (tid, criado))
        for i, (tid, uid, inicio, fim, acao, nota) in enumerate(periodos, start=1):
            con.execute(
                "INSERT INTO chamados_ticketattendance VALUES (?,?,?,?,?,?,?)",
                (i, tid, uid, inicio, fim, acao, nota),
            )
        con.commit()
        con.close()
        return caminho

    def _chamado_migrado(self, numero_int, criado_local, titulo="Chamado legado"):
        """Cria o chamado como a migracao original criou: CH-{id} e origem migrada."""
        from datetime import datetime

        criado = timezone.make_aware(datetime.strptime(criado_local, "%Y-%m-%d %H:%M:%S"))
        chamado = Chamado.objects.create(
            numero=f"CH-{numero_int:06d}",
            titulo=titulo,
            solicitante=self.solicitante,
            origem=self.ORIGEM,
            status=Chamado.STATUS_FECHADO,
        )
        # criado_em e auto_now_add: precisa de update para simular a data legada.
        Chamado.objects.filter(pk=chamado.pk).update(criado_em=criado)
        chamado.refresh_from_db()
        return chamado

    def _importar(self, caminho):
        from core.importa_atendimentos_legado import importar

        return importar(caminho)

    # ----- caso normal -----
    def test_importa_periodo_com_atendente_texto_e_duracao(self):
        chamado = self._chamado_migrado(3, "2026-02-09 17:00:50")
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[(3, 1, "2026-05-20 11:38:00", "2026-05-20 14:42:00", "stop", "Feito")],
        )
        rel = self._importar(caminho)
        self.assertEqual(rel["criados"], 1)

        periodo = AtendimentoHistorico.objects.get()
        self.assertEqual(periodo.chamado, chamado)
        self.assertEqual(periodo.atendente, self.ti)
        self.assertEqual(periodo.descricao_atividade, "Feito")
        self.assertEqual(periodo.tipo_encerramento, "stop")
        self.assertEqual(periodo.duracao.total_seconds(), 3 * 3600 + 4 * 60)
        # O banco antigo grava UTC naive: 11:38 UTC = 08:38 local (o mesmo valor
        # que a planilha preenchida a mao registra).
        self.assertEqual(timezone.localtime(periodo.iniciado_em).strftime("%d/%m/%Y %H:%M"), "20/05/2026 08:38")

    def test_marcador_tecnico_sai_da_descricao(self):
        self._chamado_migrado(3, "2026-02-09 17:00:50")
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[(3, 1, "2026-03-27 22:00:00", "2026-03-27 22:00:00", "stop",
                       "Ciclo importado do legado. [ERP-TI-CYCLE:206]")],
        )
        self._importar(caminho)
        self.assertEqual(AtendimentoHistorico.objects.get().descricao_atividade, "Ciclo importado do legado.")

    # ----- protecoes -----
    def test_periodo_sem_fim_nao_e_importado(self):
        # Um periodo sem fim seria "atendimento ativo": sujaria o Kanban e
        # bloquearia o Play do atendente.
        self._chamado_migrado(3, "2026-02-09 17:00:50")
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[(3, 1, "2026-07-15 11:47:14", None, "", "")],
        )
        rel = self._importar(caminho)
        self.assertEqual(rel["criados"], 0)
        self.assertEqual(rel["sem_fim"], 1)
        self.assertFalse(AtendimentoHistorico.objects.exists())

    def test_nao_encosta_em_chamado_recente(self):
        # CH-000800 foi criado no sistema novo: mesmo que o legado tenha um
        # ticket 800, nada e anexado a ele.
        recente = Chamado.objects.create(
            numero="CH-000800",
            titulo="Chamado novo",
            solicitante=self.solicitante,
            origem="Portal do solicitante",
            status=Chamado.STATUS_ABERTO,
        )
        caminho = self._banco_legado(
            tickets=[(800, "2026-02-09 17:00:50")],
            periodos=[(800, 1, "2026-05-20 11:38:00", "2026-05-20 12:00:00", "stop", "Feito")],
        )
        rel = self._importar(caminho)
        self.assertEqual(rel["criados"], 0)
        self.assertEqual(rel["sem_chamado"], 1)
        self.assertEqual(recente.atendimentos.count(), 0)
        recente.refresh_from_db()
        self.assertEqual(recente.status, Chamado.STATUS_ABERTO)  # intacto

    def test_data_divergente_e_pulada(self):
        # Numero bate, mas a data de criacao nao: nao e o mesmo chamado.
        self._chamado_migrado(3, "2026-02-09 17:00:50")
        caminho = self._banco_legado(
            tickets=[(3, "2025-01-01 08:00:00")],
            periodos=[(3, 1, "2026-05-20 11:38:00", "2026-05-20 12:00:00", "stop", "Feito")],
        )
        rel = self._importar(caminho)
        self.assertEqual(rel["criados"], 0)
        self.assertEqual(rel["data_divergente"], 1)

    def test_periodo_a_partir_do_primeiro_do_sistema_novo_e_ignorado(self):
        # Se os dois sistemas tiverem rodado em paralelo, nada e duplicado.
        chamado = self._chamado_migrado(3, "2026-02-09 17:00:50")
        from datetime import datetime

        corte = timezone.make_aware(datetime(2026, 7, 15, 14, 15))
        AtendimentoHistorico.objects.create(
            chamado=chamado, atendente=self.ti, iniciado_em=corte,
            finalizado_em=corte, duracao=timezone.timedelta(0), tipo_encerramento="stop",
        )
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[
                # valores em UTC, como o banco antigo grava: 12:32 UTC = 09:32
                # local (antes do corte de 14:15) e 20:00 UTC = 17:00 local (depois)
                (3, 1, "2026-07-15 12:32:00", "2026-07-15 12:40:00", "stop", "Antes do corte"),
                (3, 1, "2026-07-15 20:00:00", "2026-07-15 20:30:00", "stop", "Depois do corte"),
            ],
        )
        rel = self._importar(caminho)
        self.assertEqual(rel["criados"], 1)
        self.assertEqual(rel["apos_corte"], 1)
        self.assertTrue(
            AtendimentoHistorico.objects.filter(descricao_atividade="Antes do corte").exists()
        )
        self.assertFalse(
            AtendimentoHistorico.objects.filter(descricao_atividade="Depois do corte").exists()
        )

    def test_rodar_duas_vezes_nao_duplica(self):
        self._chamado_migrado(3, "2026-02-09 17:00:50")
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[(3, 1, "2026-05-20 11:38:00", "2026-05-20 12:00:00", "stop", "Feito")],
        )
        primeira = self._importar(caminho)
        segunda = self._importar(caminho)
        self.assertEqual(primeira["criados"], 1)
        self.assertEqual(segunda["criados"], 0)
        self.assertEqual(segunda["ja_existiam"], 1)
        self.assertEqual(AtendimentoHistorico.objects.count(), 1)

    def test_atendente_inexistente_no_sistema_novo_e_pulado(self):
        self._chamado_migrado(3, "2026-02-09 17:00:50")
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[(3, 99, "2026-05-20 11:38:00", "2026-05-20 12:00:00", "stop", "Feito")],
            usuarios=[(99, "sumiu.do.sistema")],
        )
        rel = self._importar(caminho)
        self.assertEqual(rel["criados"], 0)
        self.assertEqual(rel["sem_usuario"], 1)

    def test_nao_altera_nenhum_campo_do_chamado(self):
        chamado = self._chamado_migrado(3, "2026-02-09 17:00:50")
        antes = {
            "status": chamado.status,
            "atendente_atual_id": chamado.atendente_atual_id,
            "fechado_em": chamado.fechado_em,
            "titulo": chamado.titulo,
        }
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[(3, 1, "2026-05-20 11:38:00", "2026-05-20 12:00:00", "stop", "Feito")],
        )
        self._importar(caminho)
        chamado.refresh_from_db()
        self.assertEqual(chamado.status, antes["status"])
        self.assertEqual(chamado.atendente_atual_id, antes["atendente_atual_id"])
        self.assertEqual(chamado.fechado_em, antes["fechado_em"])
        self.assertEqual(chamado.titulo, antes["titulo"])

    def test_conferencia_sem_gravar(self):
        self._chamado_migrado(3, "2026-02-09 17:00:50")
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[(3, 1, "2026-05-20 11:38:00", "2026-05-20 12:00:00", "stop", "Feito")],
        )
        from core.importa_atendimentos_legado import importar

        rel = importar(caminho, gravar=False)
        self.assertEqual(rel["criados"], 1)
        self.assertFalse(AtendimentoHistorico.objects.exists())

    def test_periodos_importados_alimentam_a_planilha(self):
        # O ponto de todo o trabalho: os meses antigos deixam de sair vazios.
        self._chamado_migrado(3, "2026-02-09 17:00:50", titulo="Orcamento provedores de E-mail")
        caminho = self._banco_legado(
            tickets=[(3, "2026-02-09 17:00:50")],
            periodos=[(3, 1, "2026-05-20 11:38:00", "2026-05-20 14:42:00", "stop", "Feito")],
        )
        self._importar(caminho)

        self.client.force_login(self.ti)
        resp = self.client.get(reverse("atendimentos_planilha", args=[self.ti.id]), {"mes": "2026-05"})
        self.assertEqual(resp.status_code, 200)

        import io as _io

        import openpyxl

        ws = openpyxl.load_workbook(_io.BytesIO(resp.content)).active
        self.assertEqual(ws["E8"].value, "Orcamento provedores de E-mail")
        self.assertEqual(ws["H8"].value, "Feito")
        # 11:38->14:42 UTC no banco antigo = 08:38->11:42 na planilha (hora local)
        self.assertEqual(ws["B8"].value.strftime("%d/%m/%Y %H:%M"), "20/05/2026 08:38")
        self.assertEqual(ws["I8"].value.strftime("%d/%m/%Y %H:%M"), "20/05/2026 11:42")


class TemplatesLintTests(TestCase):
    """Erros de template que passam calados e aparecem na tela do usuario."""

    def test_nenhum_comentario_django_em_varias_linhas(self):
        # `{# ... #}` e comentario de UMA linha: quebrado em duas, o Django nao o
        # trata como comentario e o texto vai renderizado na pagina (ou dentro de
        # uma tag, virando lixo no HTML). Para varias linhas existe
        # `{% comment %}...{% endcomment %}`.
        from pathlib import Path

        raiz = Path(settings.BASE_DIR) / "templates"
        problemas = []
        for caminho in raiz.rglob("*.html"):
            for numero, linha in enumerate(caminho.read_text(encoding="utf-8").splitlines(), start=1):
                if "{#" in linha and "#}" not in linha:
                    problemas.append(f"{caminho.relative_to(raiz)}:{numero}: {linha.strip()[:60]}")
        self.assertEqual(problemas, [], "comentario {# #} em varias linhas: " + "; ".join(problemas))


class PausaAutomaticaTests(TestCase):
    """Pausa em lote no fim do expediente (17:45) e o complemento obrigatorio.

    Regra de uso: o que fica com o Play aberto e pausado automaticamente e nasce
    sem descricao; enquanto o atendente nao disser o que foi feito, ele nao usa
    Play, Pause nem Stop.
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x", first_name="Fabiano")
        self.outro = User.objects.create_user(username="outro.ti", password="x", first_name="Marina")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        grupo = Group.objects.get(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(grupo)
        self.outro.groups.add(grupo)

    def _chamado(self, titulo="Chamado do dia", atendente=None):
        return Chamado.objects.create(
            numero=Chamado.gerar_numero(),
            titulo=titulo,
            solicitante=self.common,
            atendente_atual=atendente or self.ti,
            status=Chamado.STATUS_EM_ATENDIMENTO,
        )

    def _play_aberto(self, chamado, hora=16, minuto=0, atendente=None):
        """Atendimento com Play em aberto (sem fim), como fica no fim do dia."""
        inicio = timezone.localtime().replace(hour=hora, minute=minuto, second=0, microsecond=0)
        return AtendimentoHistorico.objects.create(
            chamado=chamado, atendente=atendente or self.ti, iniciado_em=inicio
        )

    def _pausar(self, **opcoes):
        from io import StringIO

        from django.core.management import call_command

        saida = StringIO()
        call_command("pausar_expediente", stdout=saida, **opcoes)
        return saida.getvalue()

    # ----- o comando -----
    def test_pausa_o_play_aberto_no_horario_do_expediente(self):
        chamado = self._chamado()
        atendimento = self._play_aberto(chamado, hora=16, minuto=0)
        self._pausar()

        atendimento.refresh_from_db()
        fim = timezone.localtime(atendimento.finalizado_em)
        self.assertEqual(fim.strftime("%H:%M:%S"), "17:45:00")  # usa o corte, nao a hora do cron
        self.assertEqual(atendimento.tipo_encerramento, AtendimentoHistorico.TIPO_ENCERRAMENTO_PAUSE)
        self.assertEqual(atendimento.descricao_atividade, "")  # e o que sera complementado
        self.assertEqual(atendimento.duracao.total_seconds(), 105 * 60)  # 16:00 -> 17:45

        self.assertTrue(
            PausaAutomatica.objects.filter(atendimento=atendimento, complementado_em__isnull=True).exists()
        )
        evento = chamado.eventos.filter(tipo=ChamadoEvento.TIPO_PAUSA_AUTOMATICA).first()
        self.assertIsNotNone(evento)
        self.assertIn("17:45", evento.descricao)
        self.assertIn("Pendente de complemento", evento.descricao)

        chamado.refresh_from_db()
        self.assertEqual(chamado.status, Chamado.STATUS_ATRIBUIDO)

    def test_play_iniciado_depois_do_corte_fica_aberto(self):
        chamado = self._chamado()
        atendimento = self._play_aberto(chamado, hora=19, minuto=30)
        self._pausar()
        atendimento.refresh_from_db()
        self.assertIsNone(atendimento.finalizado_em)
        self.assertFalse(PausaAutomatica.objects.exists())

    def test_dry_run_nao_grava(self):
        atendimento = self._play_aberto(self._chamado())
        saida = self._pausar(dry_run=True)
        atendimento.refresh_from_db()
        self.assertIsNone(atendimento.finalizado_em)
        self.assertFalse(PausaAutomatica.objects.exists())
        self.assertIn("dry-run", saida)

    def test_pausa_todos_os_atendentes_de_uma_vez(self):
        self._play_aberto(self._chamado("Do Fabiano"), atendente=self.ti)
        self._play_aberto(self._chamado("Da Marina", atendente=self.outro), atendente=self.outro)
        self._pausar()
        self.assertEqual(PausaAutomatica.objects.count(), 2)
        self.assertEqual(PausaAutomatica.pendentes_de(self.ti).count(), 1)
        self.assertEqual(PausaAutomatica.pendentes_de(self.outro).count(), 1)

    def test_horario_configuravel(self):
        atendimento = self._play_aberto(self._chamado(), hora=15)
        self._pausar(hora="16:30")
        atendimento.refresh_from_db()
        self.assertEqual(timezone.localtime(atendimento.finalizado_em).strftime("%H:%M"), "16:30")

    # ----- o bloqueio -----
    def _pendencia_para_o_ti(self):
        atendimento = self._play_aberto(self._chamado("Pendente"))
        self._pausar()
        return PausaAutomatica.objects.get(atendimento=atendimento)

    def test_sem_play_enquanto_houver_pausa_pendente(self):
        self._pendencia_para_o_ti()
        outro_chamado = self._chamado("Outro chamado")
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("start_attendance"),
            data=json.dumps({"ticket_number": outro_chamado.numero}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 409)
        dados = resp.json()
        self.assertEqual(dados["pausas_pendentes"], 1)
        self.assertIn("pausado", dados["message"].lower())
        self.assertFalse(
            AtendimentoHistorico.objects.filter(chamado=outro_chamado, finalizado_em__isnull=True).exists()
        )

    def test_sem_pause_nem_stop_enquanto_houver_pausa_pendente(self):
        pausa = self._pendencia_para_o_ti()
        self.client.force_login(self.ti)
        for acao in ("pause", "stop"):
            resp = self.client.post(
                reverse("finish_attendance"),
                data=json.dumps(
                    {"ticket_number": pausa.atendimento.chamado.numero, "action": acao, "description": "algo"}
                ),
                content_type="application/json",
            )
            self.assertEqual(resp.status_code, 409, acao)
            self.assertEqual(resp.json()["pausas_pendentes"], 1)

    def test_pendencia_de_um_nao_bloqueia_o_outro(self):
        self._pendencia_para_o_ti()
        chamado = self._chamado("Da Marina", atendente=self.outro)
        self.client.force_login(self.outro)
        resp = self.client.post(
            reverse("start_attendance"),
            data=json.dumps({"ticket_number": chamado.numero}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)

    # ----- o complemento -----
    def test_complementar_grava_no_atendimento_e_libera(self):
        pausa = self._pendencia_para_o_ti()
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("pausa_complementar", args=[pausa.id]),
            data=json.dumps({"description": "Levantamento do servidor, continua amanha."}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        dados = resp.json()
        self.assertEqual(dados["restantes"], 0)
        self.assertTrue(dados["liberado"])

        pausa.refresh_from_db()
        self.assertFalse(pausa.pendente)
        self.assertEqual(pausa.complementado_por, self.ti)
        self.assertEqual(
            pausa.atendimento.descricao_atividade, "Levantamento do servidor, continua amanha."
        )
        evento = pausa.atendimento.chamado.eventos.filter(
            tipo=ChamadoEvento.TIPO_COMPLEMENTO_PAUSA
        ).first()
        self.assertIsNotNone(evento)
        self.assertIn("Complemento da pausa automatica por Fabiano", evento.descricao)
        self.assertIn("Levantamento do servidor", evento.descricao)

        resp = self.client.post(
            reverse("start_attendance"),
            data=json.dumps({"ticket_number": pausa.atendimento.chamado.numero}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)

    def test_complemento_exige_texto(self):
        pausa = self._pendencia_para_o_ti()
        self.client.force_login(self.ti)
        resp = self.client.post(
            reverse("pausa_complementar", args=[pausa.id]),
            data=json.dumps({"description": "   "}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)
        pausa.refresh_from_db()
        self.assertTrue(pausa.pendente)

    def test_nao_complementa_pausa_de_outro_atendente(self):
        pausa = self._pendencia_para_o_ti()
        self.client.force_login(self.outro)
        resp = self.client.post(
            reverse("pausa_complementar", args=[pausa.id]),
            data=json.dumps({"description": "nao e minha"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 404)

    def test_nao_complementa_duas_vezes(self):
        pausa = self._pendencia_para_o_ti()
        self.client.force_login(self.ti)
        corpo = json.dumps({"description": "feito"})
        self.assertEqual(
            self.client.post(
                reverse("pausa_complementar", args=[pausa.id]), data=corpo, content_type="application/json"
            ).status_code,
            200,
        )
        self.assertEqual(
            self.client.post(
                reverse("pausa_complementar", args=[pausa.id]), data=corpo, content_type="application/json"
            ).status_code,
            409,
        )

    def test_listagem_json_das_pendencias(self):
        pausa = self._pendencia_para_o_ti()
        self.client.force_login(self.ti)
        dados = self.client.get(reverse("pausas_pendentes")).json()
        self.assertEqual(dados["total"], 1)
        item = dados["pausas"][0]
        self.assertEqual(item["ticket_number"], pausa.atendimento.chamado.numero)
        self.assertEqual(item["fim"], "17:45")
        self.assertEqual(item["inicio"], "16:00")

    def test_usuario_comum_nao_acessa(self):
        self.client.force_login(self.common)
        self.assertEqual(self.client.get(reverse("pausas_pendentes")).status_code, 403)

    # ----- tela e planilha -----
    def test_kanban_mostra_o_aviso_pulsante(self):
        self._pendencia_para_o_ti()
        self.client.force_login(self.ti)
        html = self.client.get(reverse("tickets_dashboard")).content.decode()
        self.assertIn('id="pausaAlerta"', html)
        self.assertIn('data-pausas-pendentes="1"', html)
        self.assertIn('id="pausasPendentesModal"', html)
        self.assertIn("Play, o Pause e o Stop ficam bloqueados", html)

    def test_kanban_sem_pendencia_nao_mostra_aviso(self):
        self.client.force_login(self.ti)
        html = self.client.get(reverse("tickets_dashboard")).content.decode()
        self.assertNotIn('id="pausaAlerta"', html)
        self.assertIn('data-pausas-pendentes="0"', html)

    def test_planilha_avisa_quando_falta_o_complemento(self):
        pausa = self._pendencia_para_o_ti()
        self.client.force_login(self.ti)
        hoje = timezone.localdate()

        import io as _io

        import openpyxl

        resp = self.client.get(
            reverse("atendimentos_planilha", args=[self.ti.id]), {"mes": f"{hoje.year}-{hoje.month:02d}"}
        )
        ws = openpyxl.load_workbook(_io.BytesIO(resp.content)).active
        self.assertEqual(
            ws["H8"].value, "Pausa automatica no fim do expediente (pendente de complemento)"
        )

        pausa.complementar(descricao="Feito o levantamento", usuario=self.ti)
        resp = self.client.get(
            reverse("atendimentos_planilha", args=[self.ti.id]), {"mes": f"{hoje.year}-{hoje.month:02d}"}
        )
        ws = openpyxl.load_workbook(_io.BytesIO(resp.content)).active
        self.assertEqual(ws["H8"].value, "Feito o levantamento")

    def test_uma_linha_por_dia_na_planilha(self):
        # O ciclo real: pausa as 17:45, Play de novo no dia seguinte, pausa outra
        # vez. Cada pedaco e um periodo e sai como uma linha.
        chamado = self._chamado("Tarefa de dois dias")
        # Dias fixos do mes corrente: usar "ontem" quebraria o teste no dia 1o,
        # quando o dia anterior cai no mes passado e sai da planilha.
        agora = timezone.localtime()
        for dia_do_mes, (h1, m1, h2, m2) in ((10, (14, 0, 17, 45)), (11, (8, 0, 10, 30))):
            base = agora.replace(day=dia_do_mes)
            inicio = base.replace(hour=h1, minute=m1, second=0, microsecond=0)
            fim = base.replace(hour=h2, minute=m2, second=0, microsecond=0)
            AtendimentoHistorico.objects.create(
                chamado=chamado,
                atendente=self.ti,
                iniciado_em=inicio,
                finalizado_em=fim,
                duracao=fim - inicio,
                tipo_encerramento="pause",
                descricao_atividade=f"Parte de {h1}h",
            )
        self.client.force_login(self.ti)
        hoje = timezone.localdate()

        import io as _io

        import openpyxl

        resp = self.client.get(
            reverse("atendimentos_planilha", args=[self.ti.id]), {"mes": f"{hoje.year}-{hoje.month:02d}"}
        )
        ws = openpyxl.load_workbook(_io.BytesIO(resp.content)).active
        titulos = [ws.cell(row=r, column=5).value for r in (8, 9)]
        self.assertEqual(titulos, ["Tarefa de dois dias"] * 2)
        self.assertEqual(ws["H8"].value, "Parte de 14h")
        self.assertEqual(ws["H9"].value, "Parte de 8h")


class CasamentoNomeRamalTests(TestCase):
    """Casamento de nome de pessoa com a lista de Ramais.

    Os modulos Contatos e Kaspersky foram removidos, mas esta funcao continua no
    codigo: a planilha mensal de atendimentos usa ela para achar o setor do
    solicitante e o telefone do atendente. Os casos abaixo saem todos da base de
    producao.
    """

    def setUp(self):
        # O banco de teste ja vem com os ramais reais (seed da migration 0013).
        # Varios dos nomes usados aqui existem la, o que criaria homonimo e faria
        # a funcao (corretamente) desistir de adivinhar. Limpa para os casos
        # abaixo serem deterministicos.
        Ramal.objects.all().delete()
        ContaEmail.objects.all().delete()

    def _casar(self, nome):
        from core.views import _ramal_por_nome

        return _ramal_por_nome(nome)

    def test_nome_invertido_e_nomes_do_meio(self):
        ana = Ramal.objects.create(colaborador="Ana Gabriele", setor="Compras")
        tamara = Ramal.objects.create(colaborador="Tamara Garbuio", setor="Qualidade")
        self.assertEqual(self._casar("Gabriele Ana"), ana)                  # invertido
        self.assertEqual(self._casar("Garbuio Tamara Cristiane"), tamara)   # nome do meio a mais
        self.assertIsNone(self._casar("Fulano Que Nao Existe"))

    def test_grafia_diferente_do_sobrenome(self):
        everaldo = Ramal.objects.create(colaborador="Everaldo Vichi", setor="Qualidade")
        Ramal.objects.create(colaborador="Marcelo Costa", setor="Montagem")
        marcelo = Ramal.objects.create(colaborador="Marcelo Giannourenco", setor="Expedicao")
        self.assertEqual(self._casar("Vich Everaldo"), everaldo)
        # Entre dois "Marcelo", ganha quem tambem tem o sobrenome parecido.
        self.assertEqual(self._casar("Giamlourenco Marcelo"), marcelo)

    def test_nome_de_uma_palavra_so_exige_igualdade(self):
        portaria = Ramal.objects.create(colaborador="Portaria", setor="Portaria")
        self.assertEqual(self._casar("portaria"), portaria)
        self.assertIsNone(self._casar("porta"))

    def test_nao_casa_quando_so_o_sobrenome_bate(self):
        # Vinculo errado e pior que nenhum: a pessoa sairia do relatorio errado.
        Ramal.objects.create(colaborador="Albeni Silva", setor="Montagem Externa")
        Ramal.objects.create(colaborador="Ariadny Silva", setor="PCP")
        Ramal.objects.create(colaborador="Andre Luis", setor="Qualidade")
        self.assertIsNone(self._casar("Silva Andre"))

    def test_palavra_curta_nao_vale_como_semelhanca(self):
        Ramal.objects.create(colaborador="Ane Souza", setor="RH")
        self.assertIsNone(self._casar("Ana Souza"))

    def test_ponte_pelo_nome_completo_da_conta_de_email(self):
        # Ramal com nome curto ("Joao Leal") x GLPI "Leal Henrique": so "leal" em
        # comum. O nome completo do Workspace cobre os dois e o vinculo sai pelo
        # e-mail, que e chave exata.
        conta = ContaEmail.objects.create(
            email="henrique.leal@sidertec.com.br", primeiro_nome="Joao Henrique", sobrenome="Gomes Leal"
        )
        joao = Ramal.objects.create(
            colaborador="Joao Leal", setor="Pintura", email="henrique.leal@sidertec.com.br", conta_email=conta
        )
        self.assertEqual(self._casar("Leal Henrique"), joao)

    def test_ponte_do_email_nao_vale_com_duas_contas_possiveis(self):
        ContaEmail.objects.create(email="a@x.com", primeiro_nome="Joao Henrique", sobrenome="Gomes Leal")
        ContaEmail.objects.create(email="b@x.com", primeiro_nome="Maria Henrique", sobrenome="Leal")
        Ramal.objects.create(colaborador="Joao Leal", setor="Pintura", email="a@x.com")
        self.assertIsNone(self._casar("Leal Henrique"))

    def test_homonimos_nao_sao_adivinhados(self):
        Ramal.objects.create(colaborador="Joao Silva", setor="PCP")
        Ramal.objects.create(colaborador="Joao Silva", setor="Pintura")
        self.assertIsNone(self._casar("Joao Silva"))


class RamalKasperskyTests(TestCase):
    """Coluna "Kaspersky instalado" na lista de Ramais.

    Substitui os modulos Contatos/Kaspersky: em vez de cruzar o inventario do
    GLPI com o export do portal, o controle e um tique feito a mao na propria
    lista de pessoas.
    """

    def setUp(self):
        User = get_user_model()
        self.common = User.objects.create_user(username="comum", password="x")
        self.ti = User.objects.create_user(username="ti", password="x")
        Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)
        self.ti.groups.add(Group.objects.get(name=ATTENDANT_GROUP_NAME))
        # O banco de teste vem com os ramais do seed: limpa para os contadores
        # desta classe serem exatos.
        Ramal.objects.all().delete()
        self.ana = Ramal.objects.create(colaborador="Ana Souza", setor="RH")
        self.bruno = Ramal.objects.create(
            colaborador="Bruno Lima", setor="TI", kaspersky_instalado=True
        )

    def _toggle(self, ramal, instalado):
        return self.client.post(
            reverse("ramal_kaspersky_toggle", args=[ramal.id]),
            data=json.dumps({"instalado": instalado}),
            content_type="application/json",
        )

    def test_campo_nasce_desmarcado(self):
        self.assertFalse(Ramal.objects.create(colaborador="Novo").kaspersky_instalado)

    def test_marcar_e_desmarcar_pela_lista(self):
        self.client.force_login(self.ti)
        resp = self._toggle(self.ana, True)
        self.assertEqual(resp.status_code, 200)
        dados = resp.json()
        self.assertTrue(dados["instalado"])
        self.assertEqual(dados["com_kaspersky"], 2)  # Ana + Bruno
        self.assertEqual(dados["sem_kaspersky"], 0)
        self.ana.refresh_from_db()
        self.assertTrue(self.ana.kaspersky_instalado)

        resp = self._toggle(self.ana, False)
        self.assertFalse(resp.json()["instalado"])
        self.assertEqual(resp.json()["com_kaspersky"], 1)
        self.ana.refresh_from_db()
        self.assertFalse(self.ana.kaspersky_instalado)

    def test_usa_o_valor_enviado_e_nao_inverte_no_servidor(self):
        # Dois cliques que mandam o mesmo valor nao devem "desfazer" o tique.
        self.client.force_login(self.ti)
        self._toggle(self.ana, True)
        self._toggle(self.ana, True)
        self.ana.refresh_from_db()
        self.assertTrue(self.ana.kaspersky_instalado)

    def test_cadastro_e_edicao_levam_o_tique(self):
        self.client.force_login(self.ti)
        self.client.post(
            reverse("ramal_create"),
            {"colaborador": "Carla Dias", "setor": "PCP", "kaspersky_instalado": "on"},
        )
        carla = Ramal.objects.get(colaborador="Carla Dias")
        self.assertTrue(carla.kaspersky_instalado)

        # sem o campo no POST, o tique sai (checkbox desmarcado nao e enviado)
        self.client.post(reverse("ramal_update", args=[carla.id]), {"colaborador": "Carla Dias", "setor": "PCP"})
        carla.refresh_from_db()
        self.assertFalse(carla.kaspersky_instalado)

    def test_tela_mostra_a_coluna_o_resumo_e_o_estado(self):
        self.client.force_login(self.ti)
        resp = self.client.get(reverse("ramais_dashboard"))
        self.assertEqual(resp.context["com_kaspersky"], 1)
        self.assertEqual(resp.context["sem_kaspersky"], 1)
        html = resp.content.decode()
        self.assertIn("data-kaspersky-toggle", html)
        self.assertIn('data-kaspersky="sim"', html)   # Bruno
        self.assertIn('data-kaspersky="nao"', html)   # Ana
        # a busca acha pelos dois estados
        self.assertIn("sem kaspersky sem antivirus", html)
        self.assertIn("com kaspersky instalado antivirus", html)

    def test_usuario_comum_nao_altera(self):
        self.client.force_login(self.common)
        self.assertEqual(self._toggle(self.ana, True).status_code, 403)
        self.ana.refresh_from_db()
        self.assertFalse(self.ana.kaspersky_instalado)

    def test_ramal_inexistente_e_metodo_invalido(self):
        self.client.force_login(self.ti)
        self.assertEqual(self._toggle(Ramal(id=99999), True).status_code, 404)
        self.assertEqual(
            self.client.get(reverse("ramal_kaspersky_toggle", args=[self.ana.id])).status_code, 405
        )


class PainelTitularTests(TestCase):
    """Painel do Titular: acesso, interface do menu, usuarios, dados e trilha."""

    def setUp(self):
        User = get_user_model()
        self.admin_group, _ = Group.objects.get_or_create(name=ADMIN_GROUP_NAME)
        self.attendant_group, _ = Group.objects.get_or_create(name=ATTENDANT_GROUP_NAME)

        # O titular e sempre `fabiano.polone` (PRIMARY_ADMIN_USERNAME).
        self.titular = User.objects.create_user("fabiano.polone", password="x")
        self.titular.groups.add(self.admin_group)
        # Outro administrador (que NAO e o titular) nao deve entrar no painel.
        self.outro_admin = User.objects.create_user("maria.admin", password="x")
        self.outro_admin.groups.add(self.admin_group)
        self.comum = User.objects.create_user("joao.comum", password="x")

        self.ramal = Ramal.objects.create(colaborador="Zebedeu Teste", setor="Financeiro", ramal="1234")

    def _post(self, url, corpo=None):
        return self.client.post(url, data=json.dumps(corpo or {}), content_type="application/json")

    # ------------------------------------------------------------- acesso --
    def test_so_o_titular_entra_no_painel(self):
        for usuario in (self.outro_admin, self.comum):
            self.client.force_login(usuario)
            self.assertEqual(self.client.get(reverse("painel_titular")).status_code, 302)
            self.assertEqual(self.client.get(reverse("painel_estado")).status_code, 403)

        self.client.force_login(self.titular)
        self.assertEqual(self.client.get(reverse("painel_titular")).status_code, 200)
        self.assertEqual(self.client.get(reverse("painel_estado")).status_code, 200)

    def test_botao_do_painel_so_aparece_para_o_titular(self):
        self.client.force_login(self.outro_admin)
        self.assertNotIn("sidebar-painel", self.client.get(reverse("ramais_dashboard")).content.decode())

        self.client.force_login(self.titular)
        self.assertIn("sidebar-painel", self.client.get(reverse("ramais_dashboard")).content.decode())

    # ---------------------------------------------------------- interface --
    def test_esconder_renomear_e_mover_item_do_menu(self):
        self.client.force_login(self.titular)

        resposta = self._post(reverse("painel_interface_salvar"), {"acao": "visivel", "chave": "dicas"})
        self.assertEqual(resposta.status_code, 200)
        self.assertFalse(ItemMenuConfig.objects.get(chave="dicas").visivel)

        html = self.client.get(reverse("ramais_dashboard")).content.decode()
        self.assertNotIn("<span>Dicas</span>", html)

        self._post(reverse("painel_interface_salvar"), {"acao": "rotulo", "chave": "ramais", "valor": "Telefones"})
        html = self.client.get(reverse("ramais_dashboard")).content.decode()
        self.assertIn("<span>Telefones</span>", html)
        self.assertNotIn("<span>Ramais</span>", html)

        # Subir move o item uma posicao e grava a ordem de toda a lista.
        antes = [i["chave"] for i in itens_menu_para_painel()]
        self._post(reverse("painel_interface_salvar"), {"acao": "subir", "chave": antes[3]})
        depois = [i["chave"] for i in itens_menu_para_painel()]
        self.assertEqual(depois[2], antes[3])
        self.assertEqual(depois[3], antes[2])

        # Restaurar tudo apaga os ajustes e devolve o padrao de fabrica.
        self._post(reverse("painel_interface_salvar"), {"acao": "restaurar_tudo"})
        self.assertEqual(ItemMenuConfig.objects.count(), 0)
        self.assertEqual([i["chave"] for i in itens_menu_para_painel()], list(CHAVES_PADRAO))

    def test_primeiro_item_nao_sobe_e_chave_invalida_e_recusada(self):
        self.client.force_login(self.titular)
        primeiro = itens_menu_para_painel()[0]["chave"]
        self.assertEqual(
            self._post(reverse("painel_interface_salvar"), {"acao": "subir", "chave": primeiro}).status_code, 400
        )
        self.assertEqual(
            self._post(reverse("painel_interface_salvar"), {"acao": "visivel", "chave": "nao-existe"}).status_code, 400
        )

    # ----------------------------------------------------------- usuarios --
    def test_perfil_e_situacao_das_contas(self):
        self.client.force_login(self.titular)
        url = reverse("painel_usuario_acao", args=[self.comum.pk])

        self.assertEqual(self._post(url, {"acao": "atendente"}).status_code, 200)
        self.assertTrue(self.comum.groups.filter(name=ATTENDANT_GROUP_NAME).exists())

        self._post(url, {"acao": "admin"})
        self.assertTrue(self.comum.groups.filter(name=ADMIN_GROUP_NAME).exists())

        self._post(url, {"acao": "ativo"})
        self.comum.refresh_from_db()
        self.assertFalse(self.comum.is_active)

    def test_a_conta_do_titular_nao_pode_ser_alterada(self):
        self.client.force_login(self.titular)
        resposta = self._post(reverse("painel_usuario_acao", args=[self.titular.pk]), {"acao": "ativo"})
        self.assertEqual(resposta.status_code, 409)
        self.titular.refresh_from_db()
        self.assertTrue(self.titular.is_active)

    # -------------------------------------------------------------- dados --
    def test_listar_alterar_e_excluir_registro(self):
        self.client.force_login(self.titular)

        lista = self.client.get(reverse("painel_tabela", args=["ramais"]) + "?q=Zebedeu").json()
        self.assertEqual(lista["total"], 1)
        self.assertEqual(lista["linhas"][0]["pk"], self.ramal.pk)

        alterar = reverse("painel_registro_alterar", args=["ramais", self.ramal.pk])
        self.assertEqual(self._post(alterar, {"campo": "setor", "valor": "Compras"}).status_code, 200)
        self.ramal.refresh_from_db()
        self.assertEqual(self.ramal.setor, "Compras")

        # Booleano aceita S/N; texto que nao serve para o campo e recusado.
        self._post(alterar, {"campo": "kaspersky_instalado", "valor": "S"})
        self.ramal.refresh_from_db()
        self.assertTrue(self.ramal.kaspersky_instalado)
        self.assertEqual(self._post(alterar, {"campo": "kaspersky_instalado", "valor": "talvez"}).status_code, 400)

        # Campo automatico e campo inexistente ficam de fora.
        self.assertEqual(self._post(alterar, {"campo": "criado_em", "valor": "01/01/2026"}).status_code, 400)
        self.assertEqual(self._post(alterar, {"campo": "inventado", "valor": "x"}).status_code, 400)

        excluir = reverse("painel_registro_excluir", args=["ramais", self.ramal.pk])
        self.assertEqual(self._post(excluir).status_code, 200)
        self.assertFalse(Ramal.objects.filter(pk=self.ramal.pk).exists())

    def test_segredos_nao_saem_pelo_painel(self):
        """Nenhuma tabela pode expor senha, hash ou texto cifrado."""
        credencial = CofreCredencial.objects.create(rotulo="Roteador", usuario="admin", senha_cifrada="xxx")
        self.client.force_login(self.titular)

        detalhe = self.client.get(reverse("painel_registro", args=["cofre", credencial.pk])).json()
        nomes = [campo["nome"] for campo in detalhe["campos"]]
        self.assertNotIn("senha_cifrada", nomes)
        self.assertNotIn("xxx", json.dumps(detalhe))
        self.assertTrue(detalhe["somente_leitura"])

        # Tabela so leitura nao aceita alteracao nem exclusao.
        self.assertEqual(
            self._post(
                reverse("painel_registro_alterar", args=["cofre", credencial.pk]),
                {"campo": "rotulo", "valor": "X"},
            ).status_code,
            400,
        )
        self.assertEqual(
            self._post(reverse("painel_registro_excluir", args=["cofre", credencial.pk])).status_code, 400
        )

        for tabela in painel_dados.TABELAS:
            campos = [c.name for c in painel_dados.campos_do_modelo(tabela)]
            for suspeito in ("senha_cifrada", "senha_hash", "master_hash", "password"):
                self.assertNotIn(suspeito, campos, f"{tabela.chave} expos {suspeito}")

    # ----------------------------------------------------------- operacao --
    def test_operacao_e_trilha_de_auditoria(self):
        self.client.force_login(self.titular)

        self._post(reverse("painel_interface_salvar"), {"acao": "visivel", "chave": "dicas"})
        self._post(
            reverse("painel_registro_alterar", args=["ramais", self.ramal.pk]), {"campo": "setor", "valor": "TI"}
        )

        trilha = PainelAuditoria.objects.all()
        self.assertEqual(trilha.count(), 2)
        self.assertEqual(trilha.filter(usuario=self.titular).count(), 2)
        self.assertTrue(trilha.filter(area=PainelAuditoria.AREA_DADOS, detalhe__contains="TI").exists())

        operacao = self.client.get(reverse("painel_operacao")).json()
        self.assertEqual(len(operacao["auditoria"]), 2)
        self.assertTrue(any(l["rotulo"] == "DJANGO" for l in operacao["sistema"]))

        simulacao = self._post(reverse("painel_operacao_acao"), {"acao": "pausar_expediente_simulacao"})
        self.assertEqual(simulacao.status_code, 200)
        self.assertEqual(self._post(reverse("painel_operacao_acao"), {"acao": "inventada"}).status_code, 400)
