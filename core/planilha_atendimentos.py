"""Planilha mensal de atendimentos por atendente (modelo "Atendimentos TI Sidertec").

Gera o arquivo no mesmo formato da planilha que a TI ja preenchia a mao, a
partir do modelo versionado em `core/planilhas/modelo_atendimentos.xlsx` (todo o
layout, cores, formulas de resumo e larguras vem dele).

Regra central: **uma linha por periodo de atendimento**, ou seja, por cada
Play -> Pause/Stop (`AtendimentoHistorico`). Um mesmo chamado trabalhado em tres
dias gera tres linhas, exatamente como a planilha era preenchida a mao.

Colunas (linha 7 do modelo):

| Col | Cabecalho        | Conteudo                                              |
|-----|------------------|-------------------------------------------------------|
| A   | Tk               | vazio (nao usado nas planilhas atuais)                |
| B   | Data             | inicio do periodo (hora do Play)                      |
| C   | Contato          | solicitante do chamado                                |
| D   | Setor            | setor do solicitante, casado com a lista de Ramais    |
| E   | Notificacao      | descricao do chamado (titulo se nao houver)            |
| F   | Prioridade       | "Programada" (trabalho da TI) ou "Baixa"              |
| G   | Falha            | "N/A"                                                 |
| H   | Acao / Correcao  | o que foi feito no periodo (descricao do Pause/Stop)  |
| I   | Fechado          | fim do periodo (hora do Pause/Stop)                   |
| J   | Tempo            | duracao (fim - inicio), no formato de hora do modelo  |
| K   | Acao Eficaz      | vazio (preenchido a mao)                              |

Os numeros vao **calculados**, nao como formula. O modelo trazia `=I{n}-B{n}` no
Tempo e `COUNTIF`/`SUM` no bloco de resumo, mas o openpyxl grava formula sem valor
em cache (`<f>...</f><v/>`): o Excel abria mostrando as celulas em branco e o
grafico do resumo lia zero, mesmo com `fullCalcOnLoad` ligado no arquivo. Gravando
o valor pronto, a planilha abre certa em qualquer leitor (Excel, LibreOffice,
Sheets) sem depender de recalculo. Em troca, editar uma linha a mao nao atualiza
mais o Tempo nem os contadores sozinho.
"""

from __future__ import annotations

import io
import re
from copy import copy
from datetime import date

import openpyxl
from django.conf import settings
from django.utils import timezone

from .models import AtendimentoHistorico, Chamado

MODELO_PATH = settings.BASE_DIR / "core" / "planilhas" / "modelo_atendimentos.xlsx"

PRIMEIRA_LINHA = 8  # linha 7 e o cabecalho; os dados comecam na 8
ULTIMA_LINHA_MODELO = 152  # ate onde o modelo ja vem formatado

# Origens de chamado criadas pela propria TI: na planilha entram como
# "Programada" (trabalho planejado), nao como demanda de usuario.
ORIGENS_TI = {"Kanban TI", "Pendencia TI"}

# A planilha e preenchida com apenas dois rotulos, seguindo o criterio que a TI
# ja usava a mao: "Programada" para o trabalho da propria TI e "Baixa" para o que
# vem de usuario - independente da prioridade gravada no chamado. Conferido
# contra a planilha de 05/2026 preenchida a mao, onde chamados que o sistema
# marca como "alta"/"critica" aparecem como "Baixa".
PRIORIDADE_TI = "Programada"
PRIORIDADE_USUARIO = "Baixa"

# Metadados que a migracao do sistema antigo anexou ao fim da descricao
# ("\n\nTipo legado: programado | Falha legado: software\n[ERP-TI-ID:1]"): sao
# controle interno e nao devem aparecer na planilha.
_META_LEGADO = re.compile(r"\n{1,2}Tipo legado:.*\Z|\n?\[ERP-TI-ID:\d+\]\s*\Z", re.DOTALL)

MESES_PT = [
    "Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

FORMATO_DATA = "dd/mm/yyyy hh:mm"

# Tempo decorrido: os colchetes fazem as horas passarem de 24 em vez de dar a
# volta. O modelo trazia "[$-F400]h:mm:ss AM/PM", que mostraria um atendimento de
# 169h como "1:37:03 AM" (e ainda com um AM/PM sem sentido para duracao).
FORMATO_DURACAO = "[h]:mm:ss"


def _fim_do_mes(ano: int, mes: int) -> date:
    return date(ano + (mes // 12), (mes % 12) + 1, 1)


def periodos_do_mes(atendente, ano: int, mes: int):
    """Periodos de atendimento do atendente que **comecaram** no mes.

    O recorte usa o inicio (Play) porque e ele que define a data da linha na
    planilha; um periodo que comeca dia 31 e termina dia 1 pertence ao mes em
    que comecou. Chamados encerrados direto pelo Stop (sem Play) nao geram
    periodo e, por decisao de uso, nao entram na planilha.
    """
    inicio = date(ano, mes, 1)
    return (
        AtendimentoHistorico.objects.select_related(
            "chamado", "chamado__solicitante", "pausa_automatica"
        )
        .filter(atendente=atendente, iniciado_em__date__gte=inicio, iniciado_em__date__lt=_fim_do_mes(ano, mes))
        .order_by("iniciado_em", "id")
    )


def _prioridade_planilha(chamado: Chamado) -> str:
    """"Programada" para trabalho da propria TI; "Baixa" para demanda de usuario."""
    if (chamado.origem or "").strip() in ORIGENS_TI:
        return PRIORIDADE_TI
    # O sistema antigo tinha o tipo "programado", que veio no campo prioridade.
    if (chamado.prioridade or "").strip().lower() == "programada":
        return PRIORIDADE_TI
    solicitante = chamado.solicitante
    if solicitante is not None:
        from .permissions import is_admin_user, is_attendant_user

        if is_admin_user(solicitante) or is_attendant_user(solicitante):
            return PRIORIDADE_TI
    return PRIORIDADE_USUARIO


def _notificacao(chamado: Chamado) -> str:
    """A **descricao** do chamado: o pedido como o usuario escreveu.

    Vale para todos os chamados, migrados e novos - e o texto completo do pedido
    que interessa a quem le a planilha, nao o resumo. Cai para o titulo quando o
    chamado nao tem descricao. Os metadados que a migracao do sistema antigo
    anexou ao fim do texto ("Tipo legado: ... | Falha legado: ...",
    "[ERP-TI-ID:n]") sao removidos.
    """
    descricao = _META_LEGADO.sub("", (chamado.descricao or "").strip()).strip()
    return descricao or (chamado.titulo or "").strip()


# Bloco de resumo do modelo: a celula de contagem de cada rotulo de prioridade
# (o rotulo em si fica na coluna G da mesma linha) e o total em E2 (merge E2:E5).
LINHA_RESUMO = {"Alta": 2, "Media": 3, PRIORIDADE_USUARIO: 4, PRIORIDADE_TI: 5}
CELULA_TOTAL = "E2"


def _preencher_resumo(ws, contagem: dict, total: int) -> None:
    """Grava os contadores do cabecalho com o valor pronto (sem COUNTIF/SUM).

    O modelo traz as formulas, mas o openpyxl as grava sem valor em cache e o
    Excel abre o bloco em branco (e o grafico do resumo lê zero).
    """
    for rotulo, linha in LINHA_RESUMO.items():
        ws.cell(row=linha, column=6, value=contagem.get(rotulo, 0))
    ws[CELULA_TOTAL] = total


def _acao_correcao(periodo) -> str:
    """O que foi feito no periodo.

    Periodo pausado automaticamente no fim do expediente nasce **sem descricao**:
    ela vem do complemento que o atendente preenche no proximo acesso. Enquanto
    isso nao acontece, a planilha diz claramente que falta - em vez de sair com a
    celula vazia, que parece esquecimento de preenchimento.
    """
    texto = (periodo.descricao_atividade or "").strip()
    if texto:
        return texto
    if periodo.tipo_encerramento == "pause" and hasattr(periodo, "pausa_automatica"):
        return "Pausa automatica no fim do expediente (pendente de complemento)"
    return ""


def _copiar_estilo_linha(ws, origem: int, destino: int) -> None:
    """Replica o estilo de uma linha do modelo em outra (meses mais cheios)."""
    for col in range(1, 12):
        base = ws.cell(row=origem, column=col)
        nova = ws.cell(row=destino, column=col)
        nova._style = copy(base._style)
    if origem in ws.row_dimensions:
        ws.row_dimensions[destino].height = ws.row_dimensions[origem].height


def gerar_planilha(atendente, ano: int, mes: int, setor_por_solicitante=None, telefone_atendente: str = "") -> bytes:
    """Devolve os bytes do .xlsx do atendente para o mes informado.

    `setor_por_solicitante`: dict {user_id: setor} resolvido pela camada de view
    (vem da lista de Ramais). `telefone_atendente`: telefone que vai no
    cabecalho, ao lado do nome.
    """
    setor_por_solicitante = setor_por_solicitante or {}

    wb = openpyxl.load_workbook(MODELO_PATH)
    ws = wb.active
    ws.title = MESES_PT[mes - 1]

    nome_atendente = atendente.get_full_name() or atendente.username
    ws["A4"] = f"Atendimentos TI Sidertec - {mes:02d}/{ano}"
    ws["A5"] = f"{nome_atendente} {telefone_atendente}".strip()

    linha = PRIMEIRA_LINHA
    contagem = {PRIORIDADE_TI: 0, PRIORIDADE_USUARIO: 0}
    for periodo in periodos_do_mes(atendente, ano, mes):
        if linha > ULTIMA_LINHA_MODELO:
            _copiar_estilo_linha(ws, PRIMEIRA_LINHA, linha)

        chamado = periodo.chamado
        solicitante_id = chamado.solicitante_id

        ws.cell(row=linha, column=2, value=timezone.localtime(periodo.iniciado_em).replace(tzinfo=None))
        ws.cell(row=linha, column=2).number_format = FORMATO_DATA
        ws.cell(row=linha, column=3, value=_contato(chamado))
        ws.cell(row=linha, column=4, value=setor_por_solicitante.get(solicitante_id, ""))
        ws.cell(row=linha, column=5, value=_notificacao(chamado))
        prioridade = _prioridade_planilha(chamado)
        ws.cell(row=linha, column=6, value=prioridade)
        contagem[prioridade] = contagem.get(prioridade, 0) + 1
        ws.cell(row=linha, column=7, value="N/A")
        ws.cell(row=linha, column=8, value=_acao_correcao(periodo))

        # Periodo ainda em andamento (Play aberto): "Fechado" e "Tempo" ficam em
        # branco - nao ha duracao para informar.
        if periodo.finalizado_em:
            ws.cell(row=linha, column=9, value=timezone.localtime(periodo.finalizado_em).replace(tzinfo=None))
            ws.cell(row=linha, column=9).number_format = FORMATO_DATA
            # Duracao em dias (como o Excel guarda hora), em tempo decorrido.
            segundos = (periodo.finalizado_em - periodo.iniciado_em).total_seconds()
            ws.cell(row=linha, column=10, value=max(segundos, 0) / 86400)
            ws.cell(row=linha, column=10).number_format = FORMATO_DURACAO

        linha += 1

    _preencher_resumo(ws, contagem, total=linha - PRIMEIRA_LINHA)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _contato(chamado: Chamado) -> str:
    if chamado.solicitante_nome:
        return chamado.solicitante_nome
    if chamado.solicitante:
        return chamado.solicitante.get_full_name() or chamado.solicitante.username
    return ""


def nome_arquivo(atendente, ano: int, mes: int) -> str:
    """Mesmo padrao dos arquivos que a TI ja salva: "05-2026 - Fabiano.xlsx"."""
    nome = (atendente.first_name or atendente.get_full_name() or atendente.username).split()[0]
    return f"{mes:02d}-{ano} - {nome}.xlsx"
